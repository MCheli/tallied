"""
Unified Import API — single workflow for all document types.

Flow: Upload file → AI parsing → Review changes → Chat → Confirm
"""
import base64
import json
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import anthropic
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.dependencies import get_tenant_db, get_tenant_context, TenantContext
from app.services.audit_service import log_event
from app.models.import_session import ImportSession
from app.models.income import W2Record
from app.models.property import Mortgage
from app.models.retirement import RetirementAccount, RetirementHolding, RetirementSnapshot
from app.models.import_log import ImportLog
from app.models.account import Account
from app.models.investment import RSUGrant, VestEvent

router = APIRouter(prefix="/import", tags=["import-v1"])

# ── Context-aware prompts ──────────────────────────────────────────────────────

CONTEXT_CONFIG = {
    "income": {
        "prompt": """You are analyzing a financial document related to income (W2, pay stub, or similar).
Extract all financial values and return them as a JSON array.

Return ONLY a valid JSON array (no markdown, no explanation):
[{"field": "name", "value": "displayed value", "value_numeric": 12345.67, "category": "income", "confidence": "high", "db_field": "field_key_or_null"}]

db_field mapping:
- Gross pay / total compensation → w2_gross_pay
- Base salary / regular pay (YTD) → w2_base_salary
- RSU income / RSU gain (YTD) → w2_rsu_income
- Federal income tax → w2_federal_tax
- State income tax → w2_state_tax
- Social security tax → w2_social_security
- Medicare tax → w2_medicare
- Pre-tax 401(k) → w2_pretax_401k
- Roth 401(k) → w2_roth_401k
- Employer health insurance (Box 12 DD) → w2_employer_health
- Transit/commuter → w2_transit
- Cafeteria/Section 125 → w2_cafeteria_125

For W2 forms: use official box labels. For pay stubs: use YTD totals, not per-period amounts.
If a field doesn't match above, use null for db_field. If nothing found, return: []""",
        "expected_fields": ["w2_gross_pay", "w2_federal_tax", "w2_state_tax", "w2_social_security", "w2_medicare"],
        "guidance": "Upload a W2 form, pay stub, or other income document.",
        "examples": ["W2 form (annual tax summary)", "Pay stub (YTD salary/RSU breakdown)"],
    },
    "retirement": {
        "prompt": """You are analyzing a 401(k) or retirement account statement.
Extract key account details and return as a JSON object (NOT array):
{"plan_name": "...", "provider": "...", "total_balance": 123, "vested_balance": 123,
"roth_balance": 123, "pretax_balance": 123, "employer_match_balance": 123,
"pretax_deferral_rate_pct": 8, "roth_deferral_rate_pct": 2,
"total_employee_contributions_lifetime": 123, "total_employer_contributions_lifetime": 123,
"roth_basis": 123, "account_return_pct": 14.74, "estimated_monthly_income": 7976,
"statement_start": "YYYY-MM-DD", "statement_end": "YYYY-MM-DD",
"holdings": [{"fund_name": "...", "ticker": "...", "balance": 123, "allocation_pct": 100, "gain_loss": 123}]}
Use null for fields not found.""",
        "expected_fields": ["total_balance", "pretax_deferral_rate_pct", "roth_deferral_rate_pct"],
        "guidance": "Upload your quarterly or annual 401(k) statement.",
        "examples": ["401(k) statement from T. Rowe Price, Fidelity, Vanguard, etc."],
    },
    "property": {
        "prompt": """You are analyzing a mortgage statement.
Extract key mortgage details as a JSON array:
[{"field": "name", "value": "displayed", "value_numeric": 123, "category": "mortgage", "confidence": "high", "db_field": "key_or_null"}]

db_field mapping:
- Unpaid principal balance → mortgage_balance
- Interest rate → mortgage_rate
- Monthly payment / total payment due → mortgage_monthly_payment
- Original principal balance / original loan amount → mortgage_original_amount
- Escrow balance → mortgage_escrow_balance
- Principal portion of payment → mortgage_principal_payment
- Interest portion of payment → mortgage_interest_payment
- Escrow portion of payment → mortgage_escrow_payment
- Property address (the address of the mortgaged property) → mortgage_property_address

For the property address, set value to the full address string (e.g., "123 Main St, Springfield, MA 01101") and value_numeric to null.
If nothing found, return: []""",
        "expected_fields": ["mortgage_balance", "mortgage_rate", "mortgage_monthly_payment"],
        "guidance": "Upload your monthly mortgage statement.",
        "examples": ["Monthly mortgage statement from Chase, Wells Fargo, etc."],
    },
    "rsu": {
        "prompt": """You are analyzing an E-Trade RSU vesting schedule spreadsheet (.xlsx).
The spreadsheet has a "Restricted Stock" sheet with these row types:
- "Grant" rows: high-level grant info (Symbol, Grant Date, Granted Qty., Vested Qty., Unvested Qty., Sellable Qty., Est. Market Value, Grant Number)
- "Vest Schedule" rows: per-period vest details (Grant Number, Vest Period, Vest Date, Vested Qty., Released Qty, Total Taxes Paid, Shares Traded for taxes)
- "Tax Withholding" rows: tax detail per vest (Grant Number, Vest Period, Tax Description, Taxable Gain, Effective Tax Rate, Withholding Amount)

Extract ALL grants and their vest events. Return ONLY a valid JSON array (no markdown, no explanation):
[
  {
    "grant_number": "094907",
    "symbol": "MSFT",
    "grant_date": "2023-05-15",
    "total_shares": 200,
    "vested_shares": 150,
    "unvested_shares": 50,
    "sellable_shares": 90,
    "share_price": 95.50,
    "vest_events": [
      {
        "vest_period": 1,
        "vest_date": "2024-05-15",
        "shares": 60,
        "fmv_at_vest": 127.70,
        "shares_withheld": 22,
        "shares_released": 60,
        "total_taxes_paid": 2847.60,
        "is_actual": true
      }
    ]
  }
]

Rules:
- share_price = Est. Market Value / Granted Qty. (compute from Grant row)
- fmv_at_vest = Taxable Gain / Vested Qty. for that vest period (from Tax Withholding rows)
- total_taxes_paid = sum of all Withholding Amount values for that vest period
- is_actual = true if Vested Qty. > 0 for that vest period, false otherwise
- Convert dates to YYYY-MM-DD format
- If a field is missing or zero, use null
- Return [] if no grants found""",
        "expected_fields": ["symbol", "total_shares", "grant_number"],
        "guidance": "Upload your E-Trade RSU vesting schedule .xlsx file.",
        "examples": ["E-Trade RSU vesting schedule spreadsheet (.xlsx)"],
    },
    "general": {
        "prompt": """You are analyzing a financial document. Identify the document type and extract all relevant financial values.
Return a JSON array:
[{"field": "name", "value": "displayed", "value_numeric": 123, "category": "income|mortgage|retirement|account_balance|other", "confidence": "high|medium|low", "db_field": null}]
Try to identify if this is a W2, pay stub, mortgage statement, 401(k) statement, bank statement, or other document.
If nothing found, return: []""",
        "expected_fields": [],
        "guidance": "Upload any financial document — AI will identify the type and extract values.",
        "examples": ["W2, pay stub, mortgage statement, 401(k) statement, bank statement"],
    },
}

# W2 field mapping for applying changes
W2_FIELDS = {
    "w2_gross_pay": "gross_pay", "w2_federal_tax": "federal_tax", "w2_state_tax": "state_tax",
    "w2_social_security": "social_security", "w2_medicare": "medicare", "w2_pretax_401k": "pretax_401k",
    "w2_roth_401k": "roth_401k", "w2_cafeteria_125": "cafeteria_125", "w2_transit": "transit",
    "w2_employer_health": "employer_health", "w2_rsu_income": "rsu_income", "w2_base_salary": "base_salary",
}

MORTGAGE_FIELDS = {
    "mortgage_balance": "current_balance", "mortgage_rate": "rate",
    "mortgage_monthly_payment": "monthly_payment", "mortgage_original_amount": "original_amount",
    "mortgage_property_address": "property_address",
    "mortgage_escrow_payment": "escrow",  # Monthly escrow payment amount
}

# Fields that store text (not Decimal)
TEXT_DB_FIELDS = {"property_address"}


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.post("/upload")
def upload_and_process(
    file: UploadFile = File(...),
    context: str = Form("general"),
    db: Session = Depends(get_tenant_db),
):
    """Upload a file and start AI processing. Returns session ID."""
    if not settings.anthropic_api_key:
        raise HTTPException(status_code=500, detail="Anthropic API key not configured")
    if context not in CONTEXT_CONFIG:
        context = "general"

    session_id = str(uuid.uuid4())[:12]
    config = CONTEXT_CONFIG[context]

    # Read and encode file
    file_bytes = file.file.read()
    filename = file.filename or "upload"
    file_type = filename.rsplit(".", 1)[-1].lower() if "." in filename else "pdf"

    # Create session
    session = ImportSession(
        id=session_id, status="processing", context=context,
        filename=filename, file_type=file_type,
        chat_messages=json.dumps([]),
    )
    db.add(session)
    db.flush()

    # Send to Claude
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    try:
        if file_type == "pdf":
            # PDFs: send as document block
            pdf_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
            content_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}}
        elif file_type in ("xlsx", "xls"):
            # Excel: convert to text via openpyxl, send as text block
            import io
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                text_parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_parts.append(f"=== Sheet: {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True):
                        text_parts.append("\t".join(str(c) if c is not None else "" for c in row))
                content_block = {"type": "text", "text": f"Excel file contents ({filename}):\n" + "\n".join(text_parts)}
            except ImportError:
                content_block = {"type": "text", "text": f"Excel file uploaded: {filename} (could not parse — openpyxl not installed)"}
        else:
            # CSV and other text: send as text block
            content_block = {"type": "text", "text": f"File contents ({filename}):\n{file_bytes.decode('utf-8', errors='ignore')}"}

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16384,
            messages=[{"role": "user", "content": [content_block, {"type": "text", "text": config["prompt"]}]}],
        )
    except anthropic.APIError as e:
        session.status = "error"
        session.raw_ai_response = str(e)
        db.commit()
        raise HTTPException(status_code=502, detail=f"AI processing failed: {e}")
    except Exception as e:
        session.status = "error"
        session.raw_ai_response = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Processing error: {e}")

    try:
        raw_text = message.content[0].text.strip()
        session.raw_ai_response = raw_text

        # Parse response
        findings = _parse_ai_response(raw_text, context)
        session.parsed_findings = json.dumps(findings)

        # Build proposed changes
        proposed = _build_proposed_changes(findings, context, db, filename=filename)
        session.proposed_changes = json.dumps(proposed)

        # Identify missing fields
        if context == "retirement" and len(findings) == 1 and isinstance(findings[0], dict):
            # Retirement findings are a single object with keys as field names
            found_fields = set(findings[0].keys())
        elif context == "rsu" and len(findings) > 0 and isinstance(findings[0], dict) and "grant_number" in findings[0]:
            # RSU findings: collect field names from all grants
            found_fields = set()
            for g in findings:
                found_fields.update(k for k, v in g.items() if v is not None)
        else:
            found_fields = {f.get("db_field") for f in findings if f.get("db_field")}
        missing = [
            {"field": ef, "message": f"Expected but not found: {ef}"}
            for ef in config["expected_fields"] if ef not in found_fields
        ]
        session.missing_fields = json.dumps(missing)

        # Generate summary
        ai_summary = _generate_summary(findings, proposed, missing, filename, context)
        session.ai_summary = ai_summary
        session.status = "review"

        db.commit()
    except Exception as e:
        session.status = "error"
        db.commit()
        raise HTTPException(status_code=500, detail=f"Error parsing AI response: {e}")

    # Note: after db.commit(), SQLAlchemy expires ORM attributes.  Accessing
    # them would trigger a lazy-load on a new connection whose search_path
    # has been reset to 'public', causing "relation does not exist".  We
    # therefore return only local variables that were computed before commit.
    return {
        "session_id": session_id,
        "status": "review",
        "filename": filename,
        "context": context,
        "ai_summary": ai_summary,
        "missing_fields": missing,
        "proposed_changes": proposed,
        "findings_count": len(findings),
    }


@router.get("/{session_id}")
def get_session(session_id: str, db: Session = Depends(get_tenant_db)):
    """Get import session status and data."""
    session = db.get(ImportSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    return {
        "session_id": session.id,
        "status": session.status,
        "context": session.context,
        "filename": session.filename,
        "ai_summary": session.ai_summary,
        "missing_fields": json.loads(session.missing_fields or "[]"),
        "proposed_changes": json.loads(session.proposed_changes or "[]"),
        "chat_messages": json.loads(session.chat_messages or "[]"),
        "changes_applied": session.changes_applied,
    }


class AcceptRequest(BaseModel):
    accepted_change_ids: list[str]


@router.put("/{session_id}/accept")
def accept_changes(session_id: str, body: AcceptRequest, db: Session = Depends(get_tenant_db)):
    """Mark which proposed changes to accept."""
    session = db.get(ImportSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    session.accepted_change_ids = json.dumps(body.accepted_change_ids)
    db.commit()
    return {"accepted": len(body.accepted_change_ids)}


@router.post("/{session_id}/confirm")
def confirm_import(session_id: str, db: Session = Depends(get_tenant_db), ctx: TenantContext = Depends(get_tenant_context)):
    """Apply accepted changes to the database."""
    session = db.get(ImportSession, session_id)
    if not session or session.status != "review":
        raise HTTPException(status_code=400, detail="Session not in review state")

    proposed = json.loads(session.proposed_changes or "[]")
    accepted_ids = set(json.loads(session.accepted_change_ids or "[]"))

    # If no explicit acceptance, accept all
    if not accepted_ids:
        accepted_ids = {c["id"] for c in proposed}

    changes_applied = 0
    results = []

    try:
        for change in proposed:
            if change["id"] not in accepted_ids:
                continue

            table = change["table"]
            fields = change["fields"]

            if table == "w2_records":
                changes_applied += _apply_w2_changes(fields, change.get("tax_year"), db)
                results.append({"table": "w2_records", "fields": len(fields)})

            elif table == "mortgages":
                changes_applied += _apply_mortgage_changes(fields, db)
                results.append({"table": "mortgages", "fields": len(fields)})

            elif table == "retirement_accounts":
                changes_applied += _apply_retirement_changes(fields, db)
                results.append({"table": "retirement_accounts", "fields": len(fields)})
                # Also import holdings if present
                change_data = change.get("data", {})
                holdings = change_data.get("holdings", []) if isinstance(change_data, dict) else []
                if holdings:
                    ret_acct = db.execute(
                        select(RetirementAccount).order_by(RetirementAccount.updated_at.desc()).limit(1)
                    ).scalar_one_or_none()
                    if ret_acct:
                        h_count = _apply_retirement_holdings(holdings, ret_acct.id, db)
                        results.append({"table": "retirement_holdings", "count": h_count})

            elif table == "rsu_grants":
                data = change.get("data", {})
                changes_applied += _apply_rsu_changes(data, db)
                results.append({"table": "rsu_grants", "grant_number": data.get("grant_number")})

        session_status = "confirmed"
        session_confirmed_at = datetime.now()

        db.add(ImportLog(
            source="unified_import", capture_mode="pdf", status="confirmed",
            raw_data=json.dumps({"session_id": session_id, "filename": session.filename}),
            parsed_summary=f"Applied {changes_applied} changes",
            records_created=changes_applied,
        ))

        session.status = session_status
        session.confirmed_at = session_confirmed_at
        session.changes_applied = changes_applied
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error applying changes: {e}")

    log_event(
        action="data_import", resource_type="import_session", resource_id=session_id,
        user_id=ctx.user_id,
        details={"filename": session.filename, "changes_applied": changes_applied, "results": results},
    )

    return {"status": "confirmed", "changes_applied": changes_applied, "results": results}


class ChatRequest(BaseModel):
    message: str


@router.post("/{session_id}/chat")
def chat_with_import(session_id: str, body: ChatRequest, db: Session = Depends(get_tenant_db)):
    """Send a message to AI about the import findings."""
    session = db.get(ImportSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if not settings.anthropic_api_key:
        raise HTTPException(status_code=500, detail="API key not configured")

    # Build chat context
    messages = json.loads(session.chat_messages or "[]")
    messages.append({"role": "user", "content": body.message, "timestamp": datetime.now().isoformat()})

    proposed = json.loads(session.proposed_changes or "[]")
    findings = json.loads(session.parsed_findings or "[]")

    # Build context-specific field descriptions
    CONTEXT_FIELDS_DESC = {
        "income": """Our database tracks these income fields in the w2_records table:
- gross_pay: Total annual compensation (salary + RSU + other)
- base_salary: Regular salary (not on W2 — from pay stubs)
- rsu_income: RSU vesting income (not on W2 — from pay stubs)
- federal_tax, state_tax, social_security, medicare: Tax withholdings
- pretax_401k, roth_401k: 401(k) contributions
- cafeteria_125: FSA/cafeteria plan deductions
- transit: Commuter benefits
- employer_health: Employer health insurance contribution (Box 12 DD)""",
        "property": """Our database tracks these mortgage fields in the mortgages table:
- current_balance: Unpaid principal balance (the amount still owed)
- rate: Interest rate (stored as decimal, e.g., 0.0375 for 3.75%)
- monthly_payment: Total monthly payment (P+I+escrow)
- original_amount: Original loan amount
We do NOT track: deed records, rental income, insurance, property tax bills, or depreciation.""",
        "retirement": """Our database tracks these 401(k) fields in the retirement_accounts table:
- total_balance, vested_balance, roth_balance, pretax_balance, employer_match_balance
- pretax_deferral_rate, roth_deferral_rate (contribution percentages)
- total_employee_contributions, total_employer_contributions, roth_basis
- Holdings: fund_name, ticker, balance, allocation_pct, gain_loss""",
        "rsu": """Our database tracks RSU grants in the rsu_grants table and vest events in vest_events.""",
    }

    fields_desc = CONTEXT_FIELDS_DESC.get(session.context, "")
    changes_summary = "No database changes proposed (all values already match)." if not proposed else json.dumps(proposed[:20])

    system_prompt = f"""You are a financial data assistant helping a user import data into their personal finance app.

Document: {session.filename} (context: {session.context})

{fields_desc}

Current import status:
- AI Summary: {session.ai_summary}
- Proposed database changes: {changes_summary}
- Missing expected fields: {session.missing_fields}

IMPORTANT RULES:
1. Only talk about fields that OUR app actually tracks (listed above). Do not mention fields we don't track.
2. If no changes are proposed, explain that the data already matches what's in the database.
3. Be concise and specific. Reference actual dollar amounts and field names from the findings.
4. If the user asks to correct a value, respond naturally AND include on a SEPARATE line:
   ACTION: {{"action": "add_field", "table": "w2_records", "column": "base_salary", "value": 155000, "tax_year": 2025}}
5. When suggesting missing documents, only suggest ones that would provide data we actually track."""

    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
    try:
        # Build message history for Claude
        claude_messages = []
        for m in messages:
            claude_messages.append({"role": m["role"], "content": m["content"]})

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            system=system_prompt,
            messages=claude_messages,
        )
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=str(e))

    ai_text = response.content[0].text.strip()

    # Check for ACTION commands
    action = None
    display_text = ai_text
    if "ACTION:" in ai_text:
        parts = ai_text.split("ACTION:", 1)
        display_text = parts[0].strip()
        try:
            action_json = parts[1].strip()
            if action_json.startswith("{"):
                action = json.loads(action_json.split("\n")[0])
        except (json.JSONDecodeError, IndexError):
            pass

    messages.append({"role": "assistant", "content": display_text, "timestamp": datetime.now().isoformat(), "action": action})
    session.chat_messages = json.dumps(messages)
    db.commit()

    return {
        "response": display_text,
        "action": action,
        "messages": messages,
    }


@router.post("/{session_id}/apply-action")
def apply_chat_action(session_id: str, db: Session = Depends(get_tenant_db)):
    """Apply the most recent chat action (add/modify a proposed change)."""
    session = db.get(ImportSession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    messages = json.loads(session.chat_messages or "[]")
    # Find the last message with an action
    action = None
    for m in reversed(messages):
        if m.get("action"):
            action = m["action"]
            break

    if not action:
        return {"applied": False, "message": "No pending action found"}

    proposed = json.loads(session.proposed_changes or "[]")

    if action.get("action") == "add_field":
        # Add a new field change to proposed changes
        table = action.get("table", "w2_records")
        column = action["column"]
        value = action["value"]

        # Find existing change for this table or create new
        existing = next((c for c in proposed if c["table"] == table), None)
        if existing:
            existing["fields"].append({
                "column": column, "old_value": None, "new_value": value,
                "is_new": True, "source": "user_corrected",
            })
        else:
            proposed.append({
                "id": str(uuid.uuid4())[:8],
                "table": table, "action": "upsert",
                "tax_year": action.get("tax_year"),
                "fields": [{"column": column, "old_value": None, "new_value": value, "is_new": True, "source": "user_corrected"}],
            })

        session.proposed_changes = json.dumps(proposed)
        db.commit()
        return {"applied": True, "message": f"Added {column} = {value} to {table}"}

    if action.get("action") == "update_field":
        # Update an existing field in proposed changes
        table = action.get("table", "w2_records")
        column = action["column"]
        value = action["value"]

        updated = False
        for change in proposed:
            if change["table"] == table:
                for field in change.get("fields", []):
                    if field["column"] == column:
                        field["old_value"] = field.get("new_value")
                        field["new_value"] = value
                        field["source"] = "user_corrected"
                        updated = True
                        break
                if not updated:
                    # Field not in proposed yet — add it
                    change["fields"].append({
                        "column": column, "old_value": None, "new_value": value,
                        "is_new": True, "source": "user_corrected",
                    })
                    updated = True
                break

        if not updated:
            # No matching table change — create one
            proposed.append({
                "id": str(uuid.uuid4())[:8],
                "table": table, "action": "upsert",
                "fields": [{"column": column, "old_value": None, "new_value": value, "is_new": True, "source": "user_corrected"}],
            })

        session.proposed_changes = json.dumps(proposed)
        db.commit()
        return {"applied": True, "message": f"Updated {column} = {value} in {table}"}

    return {"applied": False, "message": "Unknown action type"}


@router.get("/config/{context}")
def get_import_config(context: str):
    """Get guidance and examples for an import context."""
    config = CONTEXT_CONFIG.get(context, CONTEXT_CONFIG["general"])
    return {
        "context": context,
        "guidance": config["guidance"],
        "examples": config["examples"],
        "expected_fields": config["expected_fields"],
    }


# ── Helpers ────────────────────────────────────────────────────────────────────

def _parse_ai_response(raw_text: str, context: str) -> list[dict]:
    """Parse Claude's response into structured findings."""
    text = raw_text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            return [parsed]  # Retirement returns an object
    except json.JSONDecodeError:
        # Try extracting array with bracket matching
        depth = 0; start = None
        for i, ch in enumerate(text):
            if ch == "[":
                if depth == 0: start = i
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0 and start is not None:
                    try:
                        return json.loads(text[start:i + 1])
                    except json.JSONDecodeError:
                        pass
    return []


def _build_proposed_changes(findings: list[dict], context: str, db: Session, filename: str = "") -> list[dict]:
    """Build proposed database changes from parsed findings."""
    changes = []

    if context == "retirement" and isinstance(findings, list) and len(findings) == 1 and isinstance(findings[0], dict) and "total_balance" in findings[0]:
        # Retirement is a single object, not array of fields
        data = findings[0]
        fields = [{"column": k, "old_value": None, "new_value": v, "is_new": True, "source": "ai_parsed"}
                  for k, v in data.items() if v is not None and k not in ("holdings", "asset_allocation", "period_contributions")]
        plan_name = data.get("plan_name", "401(k)")
        stmt_end = data.get("statement_end", "")
        record_label = f"{plan_name} — {stmt_end}" if stmt_end else plan_name
        changes.append({
            "id": str(uuid.uuid4())[:8], "table": "retirement_accounts", "action": "upsert",
            "record_label": record_label,
            "record_key": f"plan = {plan_name}",
            "fields": fields, "data": data,
        })
        return changes

    if context == "rsu" and isinstance(findings, list) and len(findings) > 0 and isinstance(findings[0], dict) and "grant_number" in findings[0]:
        # RSU findings: array of grant objects with vest_events
        for grant_data in findings:
            grant_num = str(grant_data.get("grant_number", ""))
            symbol = grant_data.get("symbol", "")
            total = grant_data.get("total_shares")
            vested = grant_data.get("vested_shares")
            unvested = grant_data.get("unvested_shares")
            vest_count = len(grant_data.get("vest_events", []))

            # Check for existing grant
            existing = db.get(RSUGrant, grant_num) if grant_num else None
            old_total = existing.total_shares if existing else None
            old_vested = existing.vested_shares if existing else None

            fields = []
            for col, val in [
                ("symbol", symbol),
                ("grant_date", grant_data.get("grant_date")),
                ("total_shares", total),
                ("vested_shares", vested),
                ("unvested_shares", unvested),
                ("sellable_shares", grant_data.get("sellable_shares")),
                ("share_price", grant_data.get("share_price")),
            ]:
                if val is not None:
                    old_val = getattr(existing, col, None) if existing else None
                    if old_val is not None:
                        old_val = float(old_val) if isinstance(old_val, Decimal) else old_val
                        if isinstance(old_val, date):
                            old_val = old_val.isoformat()
                    fields.append({
                        "column": col, "old_value": old_val, "new_value": val,
                        "is_new": existing is None, "source": "ai_parsed",
                    })

            grant_label = f"Grant {grant_num} — {symbol}"
            if total:
                grant_label += f" ({total} shares)"

            changes.append({
                "id": str(uuid.uuid4())[:8], "table": "rsu_grants", "action": "upsert",
                "record_label": grant_label,
                "record_key": f"grant_number = {grant_num}",
                "fields": fields,
                "data": grant_data,  # Full grant data including vest_events for apply step
            })

        return changes

    # Fields that are percentages (stored as decimals, AI reports as whole numbers)
    PERCENTAGE_FIELDS = {"mortgage_rate", "mortgage_escrow_balance"}
    # Fields stored as decimals in DB (rate is 0.0375, AI says 3.75)
    RATE_DB_FIELDS = {"rate"}

    # Group findings by destination table
    w2_fields = []
    mortgage_fields = []
    tax_year = None

    for f in findings:
        db_field = f.get("db_field")
        if not db_field:
            continue

        if db_field in W2_FIELDS:
            col = W2_FIELDS[db_field]
            w2 = db.execute(select(W2Record).order_by(W2Record.tax_year.desc()).limit(1)).scalar_one_or_none()
            old_val = float(getattr(w2, col, None) or 0) if w2 else None
            new_val = f.get("value_numeric")
            if new_val is not None:
                # Skip if value is unchanged (within $1 tolerance for rounding)
                if old_val is not None and abs(old_val - new_val) < 1.0:
                    continue
                w2_fields.append({
                    "column": col, "old_value": old_val, "new_value": new_val,
                    "is_new": old_val is None or old_val == 0,
                    "source": "ai_parsed", "display_name": f.get("field", col),
                    "format": "currency",
                })
            if not tax_year and w2:
                tax_year = w2.tax_year

        elif db_field in MORTGAGE_FIELDS:
            col = MORTGAGE_FIELDS[db_field]
            is_text = col in TEXT_DB_FIELDS
            mortgage = db.execute(select(Mortgage).order_by(Mortgage.updated_at.desc()).limit(1)).scalar_one_or_none()

            if is_text:
                # Text field (e.g., property_address)
                old_val = getattr(mortgage, col, None) if mortgage else None
                new_val_text = f.get("value") or f.get("value_numeric")
                if new_val_text and str(new_val_text) != str(old_val or ""):
                    mortgage_fields.append({
                        "column": col, "old_value": old_val, "new_value": str(new_val_text),
                        "is_new": old_val is None,
                        "source": "ai_parsed", "display_name": f.get("field", col),
                        "format": "text",
                    })
                continue

            old_val = float(getattr(mortgage, col, None) or 0) if mortgage else None
            new_val = f.get("value_numeric")
            if new_val is not None:
                is_rate = col in RATE_DB_FIELDS
                fmt = "percent" if is_rate else "currency"

                # Normalize rate: AI returns 3.75, DB stores 0.0375
                if is_rate and new_val > 1:
                    new_val_normalized = new_val / 100
                else:
                    new_val_normalized = new_val

                # For display: show both as the same unit
                display_old = old_val
                display_new = new_val_normalized
                if is_rate:
                    # Display as percentage (3.75%)
                    display_old = old_val * 100 if old_val else None
                    display_new = new_val if new_val > 1 else new_val * 100

                # Skip if value is unchanged
                if old_val is not None:
                    if is_rate and abs(old_val - new_val_normalized) < 0.0001:
                        continue
                    elif not is_rate and abs(old_val - new_val) < 1.0:
                        continue

                mortgage_fields.append({
                    "column": col, "old_value": display_old, "new_value": display_new,
                    "_store_value": new_val_normalized,  # actual value to write to DB
                    "is_new": old_val is None or old_val == 0,
                    "source": "ai_parsed", "display_name": f.get("field", col),
                    "format": fmt,
                })

    if w2_fields:
        # Detect tax year from filename first, then findings
        import re
        # Check filename (e.g., "2023 W2.pdf", "W2_2024.pdf")
        filename_match = re.search(r"20\d\d", filename)
        if filename_match:
            tax_year = int(filename_match.group())
        else:
            for f in findings:
                m = re.search(r"20\d\d", f.get("field", "") + f.get("value", ""))
                if m:
                    tax_year = int(m.group())
                    break

        changes.append({
            "id": str(uuid.uuid4())[:8], "table": "w2_records", "action": "upsert",
            "tax_year": tax_year,
            "record_label": f"Tax Year {tax_year}" if tax_year else "Most Recent W2",
            "record_key": f"tax_year = {tax_year}" if tax_year else None,
            "fields": w2_fields,
        })

    if mortgage_fields:
        prop_acct = db.execute(select(Account).where(Account.account_type == "real_estate").limit(1)).scalar_one_or_none()
        prop_label = prop_acct.name if prop_acct else "Primary Mortgage"
        changes.append({
            "id": str(uuid.uuid4())[:8], "table": "mortgages", "action": "upsert",
            "record_label": prop_label,
            "record_key": "id = (most recent mortgage)",
            "fields": mortgage_fields,
        })

    return changes


def _generate_summary(findings: list, proposed: list, missing: list, filename: str, context: str) -> str:
    """Generate a natural language summary of what was found."""
    parts = []
    if isinstance(findings, list) and len(findings) > 0:
        if context == "retirement" and isinstance(findings[0], dict) and "total_balance" in findings[0]:
            bal = findings[0].get("total_balance", 0)
            parts.append(f"Found retirement account data with balance ${bal:,.2f}")
        elif context == "rsu" and isinstance(findings[0], dict) and "grant_number" in findings[0]:
            total_shares = sum(g.get("total_shares", 0) or 0 for g in findings)
            total_vests = sum(len(g.get("vest_events", [])) for g in findings)
            parts.append(f"Found {len(findings)} RSU grants ({total_shares} total shares, {total_vests} vest events)")
        else:
            mapped = sum(1 for f in findings if f.get("db_field"))
            parts.append(f"Found {len(findings)} values from {filename} ({mapped} mapped to database fields)")

    for change in proposed:
        n = len(change.get("fields", []))
        new_count = sum(1 for f in change["fields"] if f.get("is_new"))
        update_count = n - new_count
        table = change["table"].replace("_", " ").title()
        if new_count and update_count:
            parts.append(f"{table}: {new_count} new fields, {update_count} updates")
        elif new_count:
            parts.append(f"{table}: {new_count} new fields")
        elif update_count:
            parts.append(f"{table}: {update_count} field updates")

    if missing:
        names = [m["field"].replace("w2_", "").replace("_", " ") for m in missing]
        parts.append(f"Not found: {', '.join(names)}")

    # If we found data but no changes are needed, say so clearly
    if findings and not proposed and not missing:
        mapped = sum(1 for f in findings if f.get("db_field")) if isinstance(findings[0], dict) and "db_field" in findings[0] else len(findings)
        if mapped > 0:
            parts.append("All values match existing data — no changes needed")

    return ". ".join(parts) if parts else "No financial data found in this document."


def _apply_w2_changes(fields: list[dict], tax_year: Optional[int], db: Session) -> int:
    """Apply W2 field changes to database."""
    year = tax_year or date.today().year
    w2 = db.execute(select(W2Record).where(W2Record.tax_year == year)).scalar_one_or_none()
    if not w2:
        w2 = W2Record(tax_year=year)
        db.add(w2)

    count = 0
    for f in fields:
        col = f["column"]
        val = f["new_value"]
        if val is not None:
            setattr(w2, col, Decimal(str(val)))
            count += 1
    db.flush()
    return count


def _apply_mortgage_changes(fields: list[dict], db: Session) -> int:
    """Apply mortgage field changes to database."""
    mortgage = db.execute(select(Mortgage).order_by(Mortgage.updated_at.desc()).limit(1)).scalar_one_or_none()
    if not mortgage:
        acct = db.execute(select(Account).where(Account.account_type == "loan_mortgage").limit(1)).scalar_one_or_none()
        if not acct:
            # Auto-create a mortgage account
            import uuid
            acct = Account(
                id=str(uuid.uuid4()),
                name="Mortgage",
                account_type="loan_mortgage",
                display_group="Home Equity",
                include_in_nw=True,
            )
            db.add(acct)
            db.flush()
        mortgage = Mortgage(account_id=acct.id)
        db.add(mortgage)

    count = 0
    for f in fields:
        col = f["column"]
        val = f.get("_store_value", f["new_value"])
        if val is not None:
            if f.get("format") == "text" or col in TEXT_DB_FIELDS:
                setattr(mortgage, col, str(val))
            else:
                setattr(mortgage, col, Decimal(str(val)))
            count += 1
    db.flush()
    return count


def _apply_retirement_changes(fields: list[dict], db: Session) -> int:
    """Apply retirement account field changes to database.

    Handles: account data, holdings, snapshots, and date-aware updates
    (older statements only create snapshots, don't overwrite current account data).
    """
    from datetime import date as date_type
    from app.models.retirement import RetirementSnapshot, RetirementHolding

    # Field name mapping: AI field name → model column name
    FIELD_MAP = {
        "total_employee_contributions_lifetime": "total_employee_contributions",
        "total_employer_contributions_lifetime": "total_employer_contributions",
    }

    # Percentage fields (AI sends as whole numbers like 14, DB stores as 0.14)
    PCT_FIELDS = {"pretax_deferral_rate_pct", "roth_deferral_rate_pct", "account_return_pct"}
    DATE_FIELDS = {"statement_start", "statement_end", "return_period_start"}

    acct = db.execute(
        select(RetirementAccount).order_by(RetirementAccount.updated_at.desc()).limit(1)
    ).scalar_one_or_none()

    is_new = acct is None
    if is_new:
        acct = RetirementAccount()
        db.add(acct)

    # Extract statement_end from fields to determine if this is a newer statement
    new_stmt_end = None
    for f in fields:
        if f["column"] == "statement_end" and f.get("new_value"):
            try:
                new_stmt_end = date_type.fromisoformat(str(f["new_value"]))
            except (ValueError, TypeError):
                pass

    # Only update account-level data if this statement is newer (or no existing date)
    is_newer = is_new or not acct.statement_end or (new_stmt_end and new_stmt_end >= acct.statement_end)

    count = 0
    for f in fields:
        col = f["column"]
        val = f.get("_store_value", f["new_value"])
        if val is None:
            continue

        # Apply field name mapping
        db_col = FIELD_MAP.get(col, col)

        if col in PCT_FIELDS:
            if is_newer:
                actual_col = col.replace("_pct", "")
                setattr(acct, actual_col, Decimal(str(val)) / 100)
                count += 1
        elif col in DATE_FIELDS:
            if is_newer:
                try:
                    setattr(acct, col, date_type.fromisoformat(str(val)))
                    count += 1
                except (ValueError, TypeError):
                    pass
        elif hasattr(acct, db_col):
            if is_newer:
                try:
                    setattr(acct, db_col, Decimal(str(val)))
                except Exception:
                    setattr(acct, db_col, str(val))
                count += 1

    db.flush()

    # Create/update balance snapshot for historical tracking
    stmt_end = new_stmt_end or acct.statement_end
    total_balance = acct.total_balance
    if stmt_end and total_balance:
        existing_snap = db.execute(
            select(RetirementSnapshot).where(
                RetirementSnapshot.retirement_account_id == acct.id,
                RetirementSnapshot.snapshot_date == stmt_end,
            )
        ).scalar_one_or_none()
        if existing_snap:
            existing_snap.balance = total_balance
            existing_snap.roth_balance = acct.roth_balance
            existing_snap.pretax_balance = acct.pretax_balance
            existing_snap.employer_balance = acct.employer_match_balance
        else:
            db.add(RetirementSnapshot(
                retirement_account_id=acct.id,
                snapshot_date=stmt_end,
                balance=total_balance,
                roth_balance=acct.roth_balance,
                pretax_balance=acct.pretax_balance,
                employer_balance=acct.employer_match_balance,
                source="statement",
            ))
        db.flush()

    # Import holdings (from the 'data' field in proposed changes)
    # This is called from confirm_import which passes fields, but holdings
    # are in the parent change's 'data' dict. We handle them separately below.

    return count


def _apply_retirement_holdings(holdings: list[dict], account_id: int, db: Session) -> int:
    """Import fund holdings for a retirement account."""
    from app.models.retirement import RetirementHolding

    if not holdings:
        return 0

    # Clear existing holdings and replace with new ones
    existing = db.execute(
        select(RetirementHolding).where(RetirementHolding.retirement_account_id == account_id)
    ).scalars().all()
    for h in existing:
        db.delete(h)

    count = 0
    for h in holdings:
        db.add(RetirementHolding(
            retirement_account_id=account_id,
            fund_name=h.get("fund_name", "Unknown Fund"),
            ticker=h.get("ticker"),
            balance=Decimal(str(h["balance"])) if h.get("balance") is not None else None,
            allocation_pct=Decimal(str(h["allocation_pct"])) / 100 if h.get("allocation_pct") is not None else None,
            gain_loss=Decimal(str(h["gain_loss"])) if h.get("gain_loss") is not None else None,
        ))
        count += 1
    db.flush()
    return count


def _apply_rsu_changes(grant_data: dict, db: Session) -> int:
    """Apply RSU grant and vest event changes to database."""
    from sqlalchemy import and_

    grant_num = str(grant_data.get("grant_number", ""))
    if not grant_num:
        return 0

    symbol = grant_data.get("symbol", "")
    company = symbol  # Use symbol as company name (e.g. "MSFT")

    # Parse grant date
    grant_date_val = None
    gd = grant_data.get("grant_date")
    if gd:
        try:
            grant_date_val = date.fromisoformat(str(gd))
        except (ValueError, TypeError):
            pass

    total_shares = grant_data.get("total_shares")
    vested_shares = grant_data.get("vested_shares")
    unvested_shares = grant_data.get("unvested_shares")
    sellable_shares = grant_data.get("sellable_shares")
    share_price = grant_data.get("share_price")

    # Upsert grant
    existing = db.get(RSUGrant, grant_num)
    if existing:
        existing.company = company
        existing.symbol = symbol
        existing.grant_date = grant_date_val
        existing.total_shares = total_shares
        existing.vested_shares = vested_shares
        existing.unvested_shares = unvested_shares
        existing.sellable_shares = sellable_shares
        if share_price is not None:
            existing.share_price = Decimal(str(share_price))
    else:
        db.add(RSUGrant(
            id=grant_num,
            company=company,
            symbol=symbol,
            grant_date=grant_date_val,
            total_shares=total_shares,
            vested_shares=vested_shares,
            unvested_shares=unvested_shares,
            sellable_shares=sellable_shares,
            share_price=Decimal(str(share_price)) if share_price else None,
        ))

    count = 1  # Grant itself

    # Process vest events
    for ve_data in grant_data.get("vest_events", []):
        vp = ve_data.get("vest_period")
        vd_str = ve_data.get("vest_date")
        if vp is None or not vd_str:
            continue

        try:
            vest_date_val = date.fromisoformat(str(vd_str))
        except (ValueError, TypeError):
            continue

        shares = ve_data.get("shares")
        fmv = ve_data.get("fmv_at_vest")
        shares_withheld = ve_data.get("shares_withheld")
        shares_released = ve_data.get("shares_released")
        total_taxes = ve_data.get("total_taxes_paid")
        is_actual = ve_data.get("is_actual", False)

        # Check for existing vest event
        existing_ve = db.execute(
            select(VestEvent).where(
                and_(VestEvent.grant_id == grant_num, VestEvent.vest_period == vp)
            )
        ).scalar_one_or_none()

        fmv_decimal = Decimal(str(fmv)) if fmv is not None else None
        taxes_decimal = Decimal(str(total_taxes)) if total_taxes is not None else None

        if existing_ve:
            existing_ve.vest_date = vest_date_val
            existing_ve.shares = shares or 0
            existing_ve.fmv_at_vest = fmv_decimal
            existing_ve.shares_withheld = shares_withheld if is_actual else None
            existing_ve.shares_released = shares_released if is_actual else None
            existing_ve.total_taxes_paid = taxes_decimal if is_actual else None
            existing_ve.is_actual = is_actual
            existing_ve.vest_price = fmv_decimal
        else:
            db.add(VestEvent(
                grant_id=grant_num,
                vest_date=vest_date_val,
                vest_period=vp,
                shares=shares or 0,
                fmv_at_vest=fmv_decimal,
                shares_withheld=shares_withheld if is_actual else None,
                shares_released=shares_released if is_actual else None,
                total_taxes_paid=taxes_decimal if is_actual else None,
                is_actual=is_actual,
                vest_price=fmv_decimal,
            ))
        count += 1

    db.flush()
    return count
