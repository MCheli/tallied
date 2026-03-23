"""
Generate comprehensive fake financial documents for test persona: Claudius Banks.

Creates realistic-looking PDFs and Excel files for all importable document types.
These are committed to the repo for testing and demo purposes.

Usage:
    python scripts/generate_test_documents.py
"""
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.colors import (
    HexColor, black, white, grey, lightgrey, darkgrey,
)
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = Path(__file__).parent.parent / "fixtures" / "test_documents"

# =============================================================================
# CLAUDIUS BANKS -- Test Persona
# =============================================================================

PERSONA = {
    "name": "CLAUDIUS BANKS",
    "first": "Claudius",
    "last": "Banks",
    "address": "456 OAK STREET",
    "city_state_zip": "SPRINGFIELD MA 01101",
    "city": "Springfield",
    "state": "MA",
    "zip": "01101",
    "ssn_last4": "7823",
    "ssn_masked": "XXX-XX-7823",
    "ein": "04-3872156",
    "employer": "MICROSOFT CORPORATION",
    "employer_address_line1": "ONE MICROSOFT WAY",
    "employer_city_state_zip": "REDMOND, WA 98052",
    "dob": "1993-04-12",
}

# Income data for 2025 (primary year)
INCOME_2025 = {
    "gross_pay": 210000.00,
    "base_salary": 145000.00,
    "rsu_income": 65000.00,
    "w2_wages": 197700.00,  # gross - pretax_401k - cafeteria_125 - transit
    "federal_tax": 38500.00,
    "state_tax": 10400.00,
    "social_security": 10918.20,
    "medicare": 3045.00,
    "pretax_401k": 10000.00,
    "roth_401k": 5000.00,
    "cafeteria_125": 2300.00,
    "transit": 900.00,
    "employer_health": 8600.00,
    "ss_wages": 168600.00,  # SS wage base for 2025
}

INCOME_2024 = {
    "gross_pay": 198750.00,
    "base_salary": 138000.00,
    "rsu_income": 60750.00,
    "w2_wages": 186510.00,
    "federal_tax": 35200.00,
    "state_tax": 9800.00,
    "social_security": 10453.20,
    "medicare": 2881.88,
    "pretax_401k": 9200.00,
    "roth_401k": 4800.00,
    "cafeteria_125": 2200.00,
    "transit": 840.00,
    "employer_health": 8200.00,
    "ss_wages": 160200.00,
}

# High earner persona (different person, for testing)
HIGH_EARNER = {
    "name": "CLAUDIUS BANKS",
    "gross_pay": 525000.00,
    "base_salary": 225000.00,
    "rsu_income": 300000.00,
    "w2_wages": 500200.00,
    "federal_tax": 142500.00,
    "state_tax": 42750.00,
    "social_security": 10453.20,
    "medicare": 7612.50,
    "pretax_401k": 23500.00,
    "roth_401k": 0.00,
    "cafeteria_125": 1300.00,
    "transit": 0.00,
    "employer_health": 14400.00,
    "ss_wages": 168600.00,
}

# Multi-state W2 data
MULTI_STATE = {
    "gross_pay": 210000.00,
    "base_salary": 145000.00,
    "rsu_income": 65000.00,
    "w2_wages": 197700.00,
    "federal_tax": 38500.00,
    "state_tax_ma": 7280.00,
    "state_tax_ca": 3120.00,
    "state_wages_ma": 147000.00,
    "state_wages_ca": 63000.00,
    "social_security": 10918.20,
    "medicare": 3045.00,
    "pretax_401k": 10000.00,
    "roth_401k": 5000.00,
    "cafeteria_125": 2300.00,
    "transit": 900.00,
    "employer_health": 8600.00,
    "ss_wages": 168600.00,
}

MORTGAGE = {
    "account_number": "7845329100",
    "property_address": "456 Oak Street, Springfield, MA 01101",
    "original_amount": 380000.00,
    "current_balance": 362450.75,
    "rate": 4.25,
    "monthly_payment": 2450.00,
    "principal": 1102.40,
    "interest": 1283.43,
    "escrow": 64.17,
    "origination_date": "06/15/2021",
    "maturity_date": "06/15/2051",
    "escrow_balance": 3215.80,
}

RETIREMENT_FIDELITY = {
    "plan_name": "MICROSOFT CORPORATION 401(K) PLAN",
    "provider": "Fidelity Investments",
    "total_balance": 185432.50,
    "vested_balance": 185432.50,
    "roth_balance": 129802.75,
    "pretax_balance": 18543.25,
    "employer_match_balance": 37086.50,
    "pretax_deferral_rate": 10,
    "roth_deferral_rate": 2,
    "roth_basis": 85000.00,
    "account_return_pct": 12.3,
    "employee_contributions_lifetime": 42000.00,
    "employer_contributions_lifetime": 12500.00,
    "estimated_monthly_income": 4520,
    "statement_start": "2025-10-01",
    "statement_end": "2025-12-31",
    "holdings": [
        {"name": "Fidelity Freedom 2055 Fund", "ticker": "FDEEX", "balance": 148346.00, "allocation": 80, "gain_loss": 18432.50},
        {"name": "Fidelity 500 Index Fund", "ticker": "FXAIX", "balance": 27828.38, "allocation": 15, "gain_loss": 4125.00},
        {"name": "Fidelity US Bond Index", "ticker": "FXNAX", "balance": 9258.12, "allocation": 5, "gain_loss": -312.50},
    ],
}

RETIREMENT_VANGUARD = {
    "plan_name": "MICROSOFT CORPORATION 401(K) PLAN",
    "provider": "Vanguard",
    "total_balance": 185432.50,
    "vested_balance": 185432.50,
    "roth_balance": 129802.75,
    "pretax_balance": 18543.25,
    "employer_match_balance": 37086.50,
    "pretax_deferral_rate": 10,
    "roth_deferral_rate": 2,
    "roth_basis": 85000.00,
    "account_return_pct": 14.74,
    "employee_contributions_lifetime": 42000.00,
    "employer_contributions_lifetime": 12500.00,
    "estimated_monthly_income": 4520,
    "statement_start": "2025-01-01",
    "statement_end": "2025-12-31",
    "holdings": [
        {"name": "Vanguard Target Retirement 2055 Fund", "ticker": "VFFVX", "balance": 185432.50, "allocation": 100, "gain_loss": 22840.00},
    ],
}

RETIREMENT_TROWEPRICE = {
    "plan_name": "MICROSOFT CORPORATION 401(K) PLAN",
    "provider": "T. Rowe Price",
    "total_balance": 185432.50,
    "vested_balance": 185432.50,
    "roth_balance": 129802.75,
    "pretax_balance": 18543.25,
    "employer_match_balance": 37086.50,
    "pretax_deferral_rate": 10,
    "roth_deferral_rate": 2,
    "roth_basis": 85000.00,
    "account_return_pct": 12.3,
    "employee_contributions_lifetime": 42000.00,
    "employer_contributions_lifetime": 12500.00,
    "estimated_monthly_income": 4520,
    "statement_start": "2025-10-01",
    "statement_end": "2025-12-31",
    "holdings": [
        {"name": "T. Rowe Price Retirement 2055 Fund", "ticker": "TRRGX", "balance": 111259.50, "allocation": 60, "gain_loss": 12300.00},
        {"name": "T. Rowe Price Blue Chip Growth", "ticker": "TRBCX", "balance": 37086.50, "allocation": 20, "gain_loss": 5840.00},
        {"name": "T. Rowe Price Equity Index 500", "ticker": "PREIX", "balance": 18543.25, "allocation": 10, "gain_loss": 2100.00},
        {"name": "T. Rowe Price Stable Value Fund", "ticker": "N/A", "balance": 18543.25, "allocation": 10, "gain_loss": 430.00},
    ],
}

RETIREMENT_MINIMAL = {
    "plan_name": "MICROSOFT CORPORATION 401(K) PLAN",
    "provider": "Fidelity Investments",
    "total_balance": 185432.50,
    "holdings": [
        {"name": "Fidelity Freedom 2055 Fund", "ticker": "FDEEX", "balance": 185432.50, "allocation": 100, "gain_loss": 0},
    ],
}

RETIREMENT_WITH_LOAN = {
    "plan_name": "MICROSOFT CORPORATION 401(K) PLAN",
    "provider": "Fidelity Investments",
    "total_balance": 185432.50,
    "vested_balance": 185432.50,
    "roth_balance": 129802.75,
    "pretax_balance": 18543.25,
    "employer_match_balance": 37086.50,
    "pretax_deferral_rate": 10,
    "roth_deferral_rate": 2,
    "roth_basis": 85000.00,
    "account_return_pct": 12.3,
    "employee_contributions_lifetime": 42000.00,
    "employer_contributions_lifetime": 12500.00,
    "estimated_monthly_income": 4520,
    "statement_start": "2025-10-01",
    "statement_end": "2025-12-31",
    "loan_balance": 15000.00,
    "loan_rate": 5.25,
    "loan_payment": 287.50,
    "loan_origination": "03/15/2025",
    "holdings": [
        {"name": "Fidelity Freedom 2055 Fund", "ticker": "FDEEX", "balance": 170432.50, "allocation": 100, "gain_loss": 5432.50},
    ],
}

RSU_GRANTS = [
    {
        "grant_number": "094907",
        "symbol": "MSFT",
        "grant_date": "2023-05-15",
        "total_shares": 200,
        "vested_shares": 100,
        "unvested_shares": 100,
        "sellable_shares": 62,
        "market_value": 84200.00,
        "vest_events": [
            {"period": 1, "date": "2024-05-15", "shares": 50, "released": 50, "taxes_paid": 6385.00, "shares_traded": 15, "taxable_gain": 21125.00, "eff_rate": 30.23},
            {"period": 2, "date": "2024-11-15", "shares": 50, "released": 50, "taxes_paid": 7200.00, "shares_traded": 16, "taxable_gain": 23750.00, "eff_rate": 30.32},
            {"period": 3, "date": "2025-05-15", "shares": 50, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
            {"period": 4, "date": "2025-11-15", "shares": 50, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
        ],
    },
    {
        "grant_number": "107234",
        "symbol": "MSFT",
        "grant_date": "2024-05-15",
        "total_shares": 250,
        "vested_shares": 0,
        "unvested_shares": 250,
        "sellable_shares": 0,
        "market_value": 105250.00,
        "vest_events": [
            {"period": 1, "date": "2025-05-15", "shares": 63, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
            {"period": 2, "date": "2025-11-15", "shares": 63, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
            {"period": 3, "date": "2026-05-15", "shares": 62, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
            {"period": 4, "date": "2026-11-15", "shares": 62, "released": 0, "taxes_paid": 0.00, "shares_traded": 0, "taxable_gain": 0.00, "eff_rate": 0.00},
        ],
    },
]


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _fmt(n):
    """Format number as currency."""
    if isinstance(n, str):
        return n
    return f"${n:,.2f}"


def _pct(n):
    """Format number as percentage."""
    return f"{n:.2f}%"


def _draw_hr(c, y, x1=0.75 * inch, x2=7.75 * inch, color=lightgrey, width=0.5):
    """Draw a horizontal rule."""
    c.setStrokeColor(color)
    c.setLineWidth(width)
    c.line(x1, y, x2, y)
    c.setStrokeColor(black)


def _draw_box(c, x, y, w, h, fill=None, stroke=lightgrey):
    """Draw a rectangle, optionally filled."""
    c.setStrokeColor(stroke)
    if fill:
        c.setFillColor(fill)
        c.rect(x, y, w, h, fill=1, stroke=1)
        c.setFillColor(black)
    else:
        c.rect(x, y, w, h, fill=0, stroke=1)


def _draw_footer(c, institution="", page=1):
    """Draw standard footer with fine print and page number."""
    c.setFont("Helvetica", 7)
    c.setFillColor(grey)
    c.drawString(0.75 * inch, 0.7 * inch,
                 f"This statement is provided for informational purposes only. Please retain for your records.")
    c.drawString(0.75 * inch, 0.55 * inch,
                 f"Questions? Contact {institution or 'your financial institution'} customer service.")
    c.drawRightString(7.75 * inch, 0.55 * inch, f"Page {page}")
    c.setFillColor(black)


def _draw_branded_header(c, institution, subtitle="", color=HexColor("#003366")):
    """Draw a branded institution header bar."""
    c.setFillColor(color)
    c.rect(0, 10.3 * inch, 8.5 * inch, 0.45 * inch, fill=1, stroke=0)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(0.75 * inch, 10.4 * inch, institution)
    if subtitle:
        c.setFont("Helvetica", 10)
        c.drawRightString(7.75 * inch, 10.42 * inch, subtitle)
    c.setFillColor(black)


# =============================================================================
# W2 GENERATORS
# =============================================================================

def _draw_w2_box(c, x, y, w, h, box_label, value, label_text):
    """Draw a single W2 box with label and value."""
    _draw_box(c, x, y - h, w, h, stroke=grey)
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(x + 2, y - 8, box_label)
    c.setFillColor(black)
    c.setFont("Helvetica", 7)
    c.drawString(x + 2, y - h + 3, label_text)
    c.setFont("Courier", 10)
    c.drawString(x + 4, y - 20, str(value))


def generate_w2_standard():
    """W2 from Microsoft with all boxes filled -- standard employer."""
    d = INCOME_2025
    filepath = OUTPUT_DIR / "income" / "w2-standard-employer.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    # Title area
    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.5 * inch, 10.5 * inch, "22222")
    c.setFont("Helvetica", 6)
    c.drawRightString(8 * inch, 10.5 * inch, "Department of the Treasury -- Internal Revenue Service")

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(4.25 * inch, 10.25 * inch, "Form W-2  Wage and Tax Statement  2025")
    c.setFont("Helvetica", 8)
    c.drawCentredString(4.25 * inch, 10.1 * inch, "Copy B -- To Be Filed With Employee's FEDERAL Tax Return")

    _draw_hr(c, 10.0 * inch, color=black, width=1)

    # Employer block (top-left)
    y = 9.8 * inch
    bw = 3.5 * inch  # box width
    bh = 0.5 * inch

    # a - Employee's SSN
    _draw_w2_box(c, 4.25 * inch, y, 3.5 * inch, 0.4 * inch, "a", PERSONA["ssn_masked"], "Employee's social security number")

    # b - Employer EIN
    _draw_w2_box(c, 0.5 * inch, y - 0.4 * inch, 3.5 * inch, 0.4 * inch, "b", PERSONA["ein"], "Employer identification number (EIN)")

    # c - Employer name/address
    _draw_box(c, 0.5 * inch, y - 1.4 * inch, 3.5 * inch, 0.6 * inch, stroke=grey)
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(0.52 * inch, y - 0.48 * inch, "c  Employer's name, address, and ZIP code")
    c.setFillColor(black)
    c.setFont("Courier", 9)
    c.drawString(0.55 * inch, y - 0.65 * inch, PERSONA["employer"])
    c.drawString(0.55 * inch, y - 0.78 * inch, PERSONA["employer_address_line1"])
    c.drawString(0.55 * inch, y - 0.91 * inch, PERSONA["employer_city_state_zip"])

    # d - Control number
    _draw_w2_box(c, 4.25 * inch, y - 0.4 * inch, 3.5 * inch, 0.4 * inch, "d", "W2-2025-001842", "Control number")

    # e/f - Employee name/address
    _draw_box(c, 4.25 * inch, y - 1.4 * inch, 3.5 * inch, 0.6 * inch, stroke=grey)
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(4.27 * inch, y - 0.48 * inch, "e  Employee's first name and initial    Last name")
    c.setFillColor(black)
    c.setFont("Courier", 9)
    c.drawString(4.3 * inch, y - 0.65 * inch, PERSONA["name"])
    c.drawString(4.3 * inch, y - 0.78 * inch, PERSONA["address"])
    c.drawString(4.3 * inch, y - 0.91 * inch, PERSONA["city_state_zip"])

    # Boxes 1-20
    y2 = y - 1.6 * inch
    bw2 = 1.75 * inch
    bh2 = 0.42 * inch

    boxes = [
        # Row 1
        ("1", _fmt(d["w2_wages"]), "Wages, tips, other compensation", 0.5 * inch),
        ("2", _fmt(d["federal_tax"]), "Federal income tax withheld", 0.5 * inch + bw2),
        ("3", _fmt(d["ss_wages"]), "Social security wages", 0.5 * inch + 2 * bw2),
        ("4", _fmt(d["social_security"]), "Social security tax withheld", 0.5 * inch + 3 * bw2),
        # Row 2
        ("5", _fmt(d["gross_pay"]), "Medicare wages and tips", 0.5 * inch),
        ("6", _fmt(d["medicare"]), "Medicare tax withheld", 0.5 * inch + bw2),
        ("7", "", "Social security tips", 0.5 * inch + 2 * bw2),
        ("8", "", "Allocated tips", 0.5 * inch + 3 * bw2),
        # Row 3
        ("9", "", "Verification code", 0.5 * inch),
        ("10", "", "Dependent care benefits", 0.5 * inch + bw2),
        ("11", "", "Nonqualified plans", 0.5 * inch + 2 * bw2),
    ]

    row = 0
    for box_num, value, label, x in boxes:
        if box_num in ("5", "9"):
            row += 1
        _draw_w2_box(c, x, y2 - row * bh2, bw2, bh2, box_num, value, label)

    # Box 12 (a-d)
    y3 = y2 - 3 * bh2
    box12_items = [
        ("12a", f"D    {_fmt(d['pretax_401k'])}", "See instructions for box 12"),
        ("12b", f"AA   {_fmt(d['roth_401k'])}", ""),
        ("12c", f"DD   {_fmt(d['employer_health'])}", ""),
        ("12d", f"W    {_fmt(d['cafeteria_125'])}", ""),
    ]
    for i, (label, value, desc) in enumerate(box12_items):
        _draw_w2_box(c, 0.5 * inch, y3 - i * 0.3 * inch, 3.5 * inch, 0.3 * inch, label, value, desc)

    # Box 13 checkboxes
    _draw_box(c, 4.25 * inch, y3 - 0.3 * inch, 3.5 * inch, 0.3 * inch, stroke=grey)
    c.setFont("Helvetica", 7)
    c.drawString(4.3 * inch, y3 - 0.2 * inch, "13  [X] Retirement plan")

    # Box 14 - Other
    _draw_box(c, 4.25 * inch, y3 - 0.9 * inch, 3.5 * inch, 0.6 * inch, stroke=grey)
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(4.27 * inch, y3 - 0.35 * inch, "14  Other")
    c.setFillColor(black)
    c.setFont("Courier", 8)
    c.drawString(4.3 * inch, y3 - 0.5 * inch, f"Transit-Salary Reduction  {_fmt(d['transit'])}")

    # State section (Boxes 15-17)
    y4 = y3 - 1.4 * inch
    _draw_box(c, 0.5 * inch, y4 - 0.5 * inch, 7.25 * inch, 0.5 * inch, stroke=grey)
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(0.52 * inch, y4 - 0.08 * inch, "15  State   Employer's state ID")
    c.drawString(3 * inch, y4 - 0.08 * inch, "16  State wages, tips, etc.")
    c.drawString(5 * inch, y4 - 0.08 * inch, "17  State income tax")
    c.setFillColor(black)
    c.setFont("Courier", 9)
    c.drawString(0.55 * inch, y4 - 0.3 * inch, f"MA   {PERSONA['ein']}")
    c.drawString(3.05 * inch, y4 - 0.3 * inch, _fmt(d["w2_wages"]))
    c.drawString(5.05 * inch, y4 - 0.3 * inch, _fmt(d["state_tax"]))

    _draw_footer(c, "Internal Revenue Service")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_w2_minimal():
    """W2 with only required boxes filled -- minimal fields."""
    d = INCOME_2024
    filepath = OUTPUT_DIR / "income" / "w2-minimal-fields.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(4.25 * inch, 10.25 * inch, "Form W-2  Wage and Tax Statement  2024")
    c.setFont("Helvetica", 8)
    c.drawCentredString(4.25 * inch, 10.1 * inch, "Copy B -- To Be Filed With Employee's FEDERAL Tax Return")

    _draw_hr(c, 10.0 * inch, color=black, width=1)

    y = 9.7 * inch
    c.setFont("Helvetica", 9)
    # Employer/Employee basic info
    c.drawString(0.75 * inch, y, f"Employer: {PERSONA['employer']}")
    c.drawString(0.75 * inch, y - 15, f"EIN: {PERSONA['ein']}")
    c.drawString(4.25 * inch, y, f"Employee: {PERSONA['name']}")
    c.drawString(4.25 * inch, y - 15, f"SSN: {PERSONA['ssn_masked']}")
    y -= 40

    # Only the essential boxes -- no RSU breakdown, no cafeteria, no transit
    fields = [
        ("Box 1", "Wages, tips, other compensation", _fmt(d["w2_wages"])),
        ("Box 2", "Federal income tax withheld", _fmt(d["federal_tax"])),
        ("Box 3", "Social security wages", _fmt(d["ss_wages"])),
        ("Box 4", "Social security tax withheld", _fmt(d["social_security"])),
        ("Box 5", "Medicare wages and tips", _fmt(d["gross_pay"])),
        ("Box 6", "Medicare tax withheld", _fmt(d["medicare"])),
        ("Box 17", "State income tax", _fmt(d["state_tax"])),
    ]

    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Wage and Tax Data")
    y -= 18

    for box, label, value in fields:
        c.setFont("Helvetica-Bold", 8)
        c.drawString(0.75 * inch, y, box)
        c.setFont("Helvetica", 8)
        c.drawString(1.25 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(6.5 * inch, y, value)
        _draw_hr(c, y - 4, x1=0.75 * inch, x2=6.5 * inch)
        y -= 18

    _draw_footer(c, "Internal Revenue Service")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_w2_multiple_states():
    """W2 with income in MA + CA -- multiple state boxes."""
    d = MULTI_STATE
    filepath = OUTPUT_DIR / "income" / "w2-multiple-states.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(4.25 * inch, 10.25 * inch, "Form W-2  Wage and Tax Statement  2025")
    c.setFont("Helvetica", 8)
    c.drawCentredString(4.25 * inch, 10.1 * inch, "Copy B -- To Be Filed With Employee's FEDERAL Tax Return")

    _draw_hr(c, 10.0 * inch, color=black, width=1)

    y = 9.7 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Employer: {PERSONA['employer']}")
    c.drawString(0.75 * inch, y - 15, f"EIN: {PERSONA['ein']}")
    c.drawString(4.25 * inch, y, f"Employee: {PERSONA['name']}")
    c.drawString(4.25 * inch, y - 15, f"SSN: {PERSONA['ssn_masked']}")
    y -= 45

    # Federal boxes
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Federal Wage and Tax Data")
    y -= 18

    fed_fields = [
        ("1", "Wages, tips, other compensation", _fmt(d["w2_wages"])),
        ("2", "Federal income tax withheld", _fmt(d["federal_tax"])),
        ("3", "Social security wages", _fmt(d["ss_wages"])),
        ("4", "Social security tax withheld", _fmt(d["social_security"])),
        ("5", "Medicare wages and tips", _fmt(d["gross_pay"])),
        ("6", "Medicare tax withheld", _fmt(d["medicare"])),
        ("12a", "Code D (401k pretax)", _fmt(d["pretax_401k"])),
        ("12b", "Code AA (Roth 401k)", _fmt(d["roth_401k"])),
        ("12c", "Code DD (Employer health)", _fmt(d["employer_health"])),
        ("14", "Transit-Salary Reduction", _fmt(d["transit"])),
    ]

    for box, label, value in fed_fields:
        c.setFont("Helvetica-Bold", 8)
        c.drawString(0.75 * inch, y, f"Box {box}")
        c.setFont("Helvetica", 8)
        c.drawString(1.5 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(6.5 * inch, y, value)
        y -= 16

    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "State Tax Information")
    y -= 18

    # State 1: MA
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 15")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State: MA    Employer State ID: {PERSONA['ein']}")
    y -= 16
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 16")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State wages: {_fmt(d['state_wages_ma'])}")
    y -= 16
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 17")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State income tax: {_fmt(d['state_tax_ma'])}")
    y -= 22

    # State 2: CA
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 15")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State: CA    Employer State ID: 91-1234567")
    y -= 16
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 16")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State wages: {_fmt(d['state_wages_ca'])}")
    y -= 16
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Box 17")
    c.setFont("Helvetica", 8)
    c.drawString(1.5 * inch, y, f"State income tax: {_fmt(d['state_tax_ca'])}")

    _draw_footer(c, "Internal Revenue Service")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_w2_high_earner():
    """W2 with very high income, all fields maxed out."""
    d = HIGH_EARNER
    filepath = OUTPUT_DIR / "income" / "w2-high-earner.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(4.25 * inch, 10.25 * inch, "Form W-2  Wage and Tax Statement  2025")
    c.setFont("Helvetica", 8)
    c.drawCentredString(4.25 * inch, 10.1 * inch, "Copy B -- To Be Filed With Employee's FEDERAL Tax Return")

    _draw_hr(c, 10.0 * inch, color=black, width=1)

    y = 9.7 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Employer: {PERSONA['employer']}")
    c.drawString(0.75 * inch, y - 15, f"EIN: {PERSONA['ein']}")
    c.drawString(4.25 * inch, y, f"Employee: {PERSONA['name']}")
    c.drawString(4.25 * inch, y - 15, f"SSN: {PERSONA['ssn_masked']}")
    y -= 45

    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Wage and Tax Data")
    y -= 18

    fields = [
        ("1", "Wages, tips, other compensation", _fmt(d["w2_wages"])),
        ("2", "Federal income tax withheld", _fmt(d["federal_tax"])),
        ("3", "Social security wages", _fmt(d["ss_wages"])),
        ("4", "Social security tax withheld", _fmt(d["social_security"])),
        ("5", "Medicare wages and tips", _fmt(d["gross_pay"])),
        ("6", "Medicare tax withheld", _fmt(d["medicare"])),
        ("12a", "Code D (401k pretax)", _fmt(d["pretax_401k"])),
        ("12c", "Code DD (Employer health)", _fmt(d["employer_health"])),
        ("14", "Cafeteria 125", _fmt(d["cafeteria_125"])),
        ("15-17", f"State: MA    State wages: {_fmt(d['w2_wages'])}    State tax: {_fmt(d['state_tax'])}", ""),
    ]

    for box, label, value in fields:
        c.setFont("Helvetica-Bold", 8)
        c.drawString(0.75 * inch, y, f"Box {box}")
        c.setFont("Helvetica", 8)
        c.drawString(1.5 * inch, y, label)
        if value:
            c.setFont("Courier", 9)
            c.drawRightString(6.5 * inch, y, value)
        y -= 16

    y -= 10
    c.setFont("Helvetica", 7)
    c.setFillColor(grey)
    c.drawString(0.75 * inch, y, "Note: Box 5 Medicare wages exceed Box 3 SS wages due to the Social Security wage base cap ($168,600 for 2025).")
    c.drawString(0.75 * inch, y - 10, "Additional Medicare Tax of 0.9% applies to wages exceeding $200,000.")
    c.setFillColor(black)

    _draw_footer(c, "Internal Revenue Service")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_paystub_adp():
    """ADP-style pay stub with YTD totals, multi-column layout."""
    d = INCOME_2025
    filepath = OUTPUT_DIR / "income" / "paystub-adp.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    # ADP branded header
    _draw_branded_header(c, "ADP", "Earnings Statement", color=HexColor("#D62B1F"))

    y = 10.1 * inch

    # Company / Employee info
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.75 * inch, y, "Company:")
    c.setFont("Helvetica", 9)
    c.drawString(1.6 * inch, y, PERSONA["employer"])
    c.drawString(1.6 * inch, y - 12, PERSONA["employer_address_line1"])
    c.drawString(1.6 * inch, y - 24, PERSONA["employer_city_state_zip"])

    c.setFont("Helvetica-Bold", 9)
    c.drawString(4.5 * inch, y, "Employee:")
    c.setFont("Helvetica", 9)
    c.drawString(5.3 * inch, y, PERSONA["name"])
    c.drawString(5.3 * inch, y - 12, f"{PERSONA['address']}")
    c.drawString(5.3 * inch, y - 24, PERSONA["city_state_zip"])
    y -= 45

    # Pay period info
    _draw_box(c, 0.75 * inch, y - 0.5 * inch, 6.75 * inch, 0.5 * inch, fill=HexColor("#F5F5F5"))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.8 * inch, y - 0.15 * inch, "Pay Date: 12/19/2025")
    c.drawString(2.5 * inch, y - 0.15 * inch, "Period: 12/07/2025 - 12/20/2025")
    c.drawString(5 * inch, y - 0.15 * inch, "Pay Frequency: Bi-Weekly")
    c.drawString(0.8 * inch, y - 0.35 * inch, f"Employee ID: E-482910")
    c.drawString(2.5 * inch, y - 0.35 * inch, f"Department: Cloud & AI")
    c.drawString(5 * inch, y - 0.35 * inch, f"Federal Filing Status: Single")
    y -= 0.7 * inch

    # EARNINGS table
    _draw_hr(c, y, color=HexColor("#D62B1F"), width=1.5)
    y -= 5
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.75 * inch, y, "EARNINGS")
    c.drawString(3.5 * inch, y, "Rate")
    c.drawString(4.5 * inch, y, "Hours")
    c.drawString(5.3 * inch, y, "This Period")
    c.drawRightString(7.5 * inch, y, "Year to Date")
    y -= 5
    _draw_hr(c, y)
    y -= 12

    earnings = [
        ("Regular", "5,576.92", "80.00", "5,576.92", _fmt(d["base_salary"])),
        ("RSU Gain", "", "", "", _fmt(d["rsu_income"])),
    ]

    c.setFont("Helvetica", 8)
    for desc, rate, hours, period, ytd in earnings:
        c.drawString(0.75 * inch, y, desc)
        c.drawString(3.5 * inch, y, rate)
        c.drawString(4.5 * inch, y, hours)
        c.drawString(5.3 * inch, y, period)
        c.drawRightString(7.5 * inch, y, ytd)
        y -= 14

    _draw_hr(c, y + 2)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y - 5, "Gross Pay")
    c.drawString(5.3 * inch, y - 5, "5,576.92")
    c.drawRightString(7.5 * inch, y - 5, _fmt(d["gross_pay"]))
    y -= 25

    # DEDUCTIONS table
    _draw_hr(c, y, color=HexColor("#D62B1F"), width=1.5)
    y -= 5
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.75 * inch, y, "DEDUCTIONS")
    c.drawString(4.5 * inch, y, "Statutory")
    c.drawString(5.3 * inch, y, "This Period")
    c.drawRightString(7.5 * inch, y, "Year to Date")
    y -= 5
    _draw_hr(c, y)
    y -= 12

    deductions = [
        ("Federal Income Tax", "Yes", "1,480.77", _fmt(d["federal_tax"])),
        ("Social Security Tax", "Yes", "419.93", _fmt(d["social_security"])),
        ("Medicare Tax", "Yes", "117.12", _fmt(d["medicare"])),
        ("MA State Income Tax", "Yes", "400.00", _fmt(d["state_tax"])),
        ("401(k) Pre-Tax", "No", "384.62", _fmt(d["pretax_401k"])),
        ("401(k) Roth After-Tax", "No", "192.31", _fmt(d["roth_401k"])),
        ("Dental + Medical (Sec 125)", "No", "88.46", _fmt(d["cafeteria_125"])),
        ("Pretax Commuter/Transit", "No", "34.62", _fmt(d["transit"])),
    ]

    c.setFont("Helvetica", 8)
    for desc, stat, period, ytd in deductions:
        c.drawString(0.75 * inch, y, desc)
        c.drawString(4.5 * inch, y, stat)
        c.drawString(5.3 * inch, y, period)
        c.drawRightString(7.5 * inch, y, ytd)
        y -= 14

    # Net pay
    y -= 10
    _draw_hr(c, y, color=HexColor("#D62B1F"), width=1.5)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "NET PAY")
    c.drawString(5.3 * inch, y, "$2,459.09")
    c.drawRightString(7.5 * inch, y, "$119,964.72")

    # Employer-paid benefits (irrelevant section)
    y -= 35
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.75 * inch, y, "EMPLOYER-PAID BENEFITS (Not deducted from your pay)")
    y -= 5
    _draw_hr(c, y)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, "Employer Health Insurance (Box 12 DD)")
    c.drawRightString(7.5 * inch, y, _fmt(d["employer_health"]))

    _draw_footer(c, "ADP, LLC")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_paystub_workday():
    """Workday-style pay stub -- modern layout, different format."""
    d = INCOME_2025
    filepath = OUTPUT_DIR / "income" / "paystub-workday.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    # Workday branded header
    _draw_branded_header(c, "Workday", "Total Rewards Statement", color=HexColor("#FF6600"))

    y = 10.0 * inch

    # Employee info in a clean card layout
    _draw_box(c, 0.75 * inch, y - 0.8 * inch, 6.75 * inch, 0.8 * inch, fill=HexColor("#FAFAFA"), stroke=HexColor("#DDDDDD"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.9 * inch, y - 0.15 * inch, PERSONA["name"])
    c.setFont("Helvetica", 8)
    c.drawString(0.9 * inch, y - 0.3 * inch, f"Employee ID: E-482910  |  Department: Cloud & AI  |  Location: Springfield, MA")
    c.drawString(0.9 * inch, y - 0.45 * inch, f"Pay Group: Bi-Weekly  |  Pay Date: December 19, 2025  |  Period End: December 20, 2025")
    c.drawString(0.9 * inch, y - 0.6 * inch, f"Manager: Jane Smith  |  Position: Senior Software Engineer (Level 63)")
    y -= 1.0 * inch

    # Summary boxes at top
    box_w = 2.1 * inch
    box_h = 0.5 * inch
    summary_items = [
        ("YTD Gross", _fmt(d["gross_pay"]), HexColor("#003366")),
        ("YTD Taxes", _fmt(d["federal_tax"] + d["state_tax"] + d["social_security"] + d["medicare"]), HexColor("#8B0000")),
        ("YTD Net", "$119,964.72", HexColor("#006400")),
    ]
    for i, (label, value, color) in enumerate(summary_items):
        x = 0.75 * inch + i * (box_w + 0.2 * inch)
        _draw_box(c, x, y - box_h, box_w, box_h, fill=color)
        c.setFillColor(white)
        c.setFont("Helvetica", 7)
        c.drawString(x + 5, y - 0.15 * inch, label)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x + 5, y - 0.4 * inch, value)
        c.setFillColor(black)
    y -= box_h + 0.25 * inch

    # Earnings detail
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(HexColor("#FF6600"))
    c.drawString(0.75 * inch, y, "EARNINGS DETAIL")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#FF6600"), width=1)
    y -= 14

    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.75 * inch, y, "Description")
    c.drawRightString(4 * inch, y, "Current")
    c.drawRightString(5.5 * inch, y, "YTD")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    earnings = [
        ("Base Salary (Bi-Weekly)", "$5,576.92", _fmt(d["base_salary"])),
        ("RSU Vesting Income", "$0.00", _fmt(d["rsu_income"])),
    ]
    c.setFont("Helvetica", 8)
    for desc, current, ytd in earnings:
        c.drawString(0.75 * inch, y, desc)
        c.drawRightString(4 * inch, y, current)
        c.drawRightString(5.5 * inch, y, ytd)
        y -= 14

    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Total Gross")
    c.drawRightString(4 * inch, y, "$5,576.92")
    c.drawRightString(5.5 * inch, y, _fmt(d["gross_pay"]))
    y -= 25

    # Tax withholdings
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(HexColor("#FF6600"))
    c.drawString(0.75 * inch, y, "TAX WITHHOLDINGS")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#FF6600"), width=1)
    y -= 14

    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.75 * inch, y, "Tax Type")
    c.drawRightString(5.5 * inch, y, "YTD")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    taxes = [
        ("Federal Income Tax", _fmt(d["federal_tax"])),
        ("Social Security (OASDI)", _fmt(d["social_security"])),
        ("Medicare", _fmt(d["medicare"])),
        ("MA State Income Tax", _fmt(d["state_tax"])),
    ]
    c.setFont("Helvetica", 8)
    for desc, ytd in taxes:
        c.drawString(0.75 * inch, y, desc)
        c.drawRightString(5.5 * inch, y, ytd)
        y -= 14

    # Pre-tax deductions
    y -= 10
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(HexColor("#FF6600"))
    c.drawString(0.75 * inch, y, "PRE-TAX DEDUCTIONS")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#FF6600"), width=1)
    y -= 14

    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.75 * inch, y, "Benefit")
    c.drawRightString(5.5 * inch, y, "YTD")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    pretax = [
        ("401(k) Pre-Tax Contribution", _fmt(d["pretax_401k"])),
        ("401(k) Roth Contribution", _fmt(d["roth_401k"])),
        ("Medical/Dental (Section 125)", _fmt(d["cafeteria_125"])),
        ("Commuter Benefits", _fmt(d["transit"])),
    ]
    c.setFont("Helvetica", 8)
    for desc, ytd in pretax:
        c.drawString(0.75 * inch, y, desc)
        c.drawRightString(5.5 * inch, y, ytd)
        y -= 14

    # Employer contributions (visual noise)
    y -= 10
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(HexColor("#FF6600"))
    c.drawString(0.75 * inch, y, "EMPLOYER CONTRIBUTIONS (Informational)")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#FF6600"), width=1)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, "Employer Health Insurance (Box 12 DD)")
    c.drawRightString(5.5 * inch, y, _fmt(d["employer_health"]))

    _draw_footer(c, "Workday, Inc.")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


# =============================================================================
# RETIREMENT (401K) GENERATORS
# =============================================================================

def _generate_401k_common(c, r, y, provider_color):
    """Common 401k statement sections shared across providers."""
    # Account overview
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(provider_color)
    c.drawString(0.75 * inch, y, "Account Summary")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=provider_color, width=1)
    y -= 14

    summary_fields = [
        ("Ending Balance", _fmt(r["total_balance"])),
        ("Vested Balance", _fmt(r.get("vested_balance", r["total_balance"]))),
    ]
    if "account_return_pct" in r:
        summary_fields.append(("Your Account Return", _pct(r["account_return_pct"])))
    if "estimated_monthly_income" in r:
        summary_fields.append(("Estimated Monthly Income at Retirement", _fmt(r["estimated_monthly_income"])))

    for label, value in summary_fields:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(5.5 * inch, y, value)
        y -= 14

    return y


def generate_401k_fidelity():
    """Fidelity quarterly 401k statement with multiple funds."""
    r = RETIREMENT_FIDELITY
    filepath = OUTPUT_DIR / "retirement" / "401k-fidelity-quarterly.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)
    fid_green = HexColor("#4A8C2A")

    _draw_branded_header(c, "Fidelity Investments", "Quarterly Statement", color=fid_green)

    y = 10.0 * inch
    # Plan/employee info
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Plan: {r['plan_name']}")
    c.drawString(0.75 * inch, y - 12, f"Participant: {PERSONA['name']}")
    c.drawString(0.75 * inch, y - 24, f"SSN: {PERSONA['ssn_masked']}")
    c.drawString(4.5 * inch, y, f"Statement Period: {r['statement_start']} to {r['statement_end']}")
    c.drawString(4.5 * inch, y - 12, f"Account Number: Z12-345678")
    y -= 45

    y = _generate_401k_common(c, r, y, fid_green)

    # Balance by type
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Balance by Contribution Type")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14

    type_fields = [
        ("Roth Contributions", _fmt(r["roth_balance"])),
        ("Salary Deferral (Pre-Tax)", _fmt(r["pretax_balance"])),
        ("Employer Match", _fmt(r["employer_match_balance"])),
    ]
    for label, value in type_fields:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(5.5 * inch, y, value)
        y -= 14

    # Lifetime contributions
    y -= 10
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, "Total Employee Contributions (since Fidelity)")
    c.drawRightString(5.5 * inch, y, _fmt(r["employee_contributions_lifetime"]))
    y -= 14
    c.drawString(0.75 * inch, y, "Total Employer Contributions (since Fidelity)")
    c.drawRightString(5.5 * inch, y, _fmt(r["employer_contributions_lifetime"]))
    y -= 14
    c.drawString(0.75 * inch, y, "Roth Basis")
    c.drawRightString(5.5 * inch, y, _fmt(r["roth_basis"]))

    # Deferral rates
    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Current Contribution Rates")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"Pre-Tax Deferral: {r['pretax_deferral_rate']}%")
    y -= 14
    c.drawString(0.75 * inch, y, f"Roth Deferral: {r['roth_deferral_rate']}%")

    # Holdings
    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Investment Holdings")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14

    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.75 * inch, y, "Fund Name")
    c.drawString(3.5 * inch, y, "Ticker")
    c.drawRightString(5 * inch, y, "Balance")
    c.drawRightString(6 * inch, y, "Alloc %")
    c.drawRightString(7.5 * inch, y, "Gain/Loss")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    c.setFont("Helvetica", 8)
    for h in r["holdings"]:
        c.drawString(0.75 * inch, y, h["name"])
        c.drawString(3.5 * inch, y, h["ticker"])
        c.drawRightString(5 * inch, y, _fmt(h["balance"]))
        c.drawRightString(6 * inch, y, f"{h['allocation']}%")
        c.drawRightString(7.5 * inch, y, _fmt(h["gain_loss"]))
        y -= 14

    # Marketing section (visual noise)
    y -= 20
    _draw_box(c, 0.75 * inch, y - 0.6 * inch, 6.75 * inch, 0.6 * inch, fill=HexColor("#F0F7E8"))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.9 * inch, y - 0.15 * inch, "Stay on track with Fidelity Planning & Guidance")
    c.setFont("Helvetica", 7)
    c.drawString(0.9 * inch, y - 0.3 * inch, "Visit NetBenefits.Fidelity.com to review your retirement readiness score,")
    c.drawString(0.9 * inch, y - 0.42 * inch, "explore investment options, or schedule a complimentary planning session.")

    _draw_footer(c, "Fidelity Investments")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_401k_vanguard():
    """Vanguard annual 401k statement with single target date fund."""
    r = RETIREMENT_VANGUARD
    filepath = OUTPUT_DIR / "retirement" / "401k-vanguard-annual.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)
    vg_red = HexColor("#96262C")

    _draw_branded_header(c, "Vanguard", "Annual Account Statement 2025", color=vg_red)

    y = 10.0 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Plan: {r['plan_name']}")
    c.drawString(0.75 * inch, y - 12, f"Participant: {PERSONA['name']}")
    c.drawString(4.5 * inch, y, f"Period: {r['statement_start']} through {r['statement_end']}")
    y -= 40

    y = _generate_401k_common(c, r, y, vg_red)

    # Balance by type
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(vg_red)
    c.drawString(0.75 * inch, y, "Account Balance Detail")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=vg_red, width=1)
    y -= 14

    for label, val in [
        ("Roth Balance", _fmt(r["roth_balance"])),
        ("Pre-Tax Balance", _fmt(r["pretax_balance"])),
        ("Employer Match", _fmt(r["employer_match_balance"])),
        ("Total Employee Contributions (lifetime)", _fmt(r["employee_contributions_lifetime"])),
        ("Total Employer Contributions (lifetime)", _fmt(r["employer_contributions_lifetime"])),
        ("Roth Basis", _fmt(r["roth_basis"])),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.drawRightString(5.5 * inch, y, val)
        y -= 14

    # Rates
    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(vg_red)
    c.drawString(0.75 * inch, y, "Contribution Elections")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=vg_red, width=1)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"Pre-Tax Deferral Rate: {r['pretax_deferral_rate']}%")
    y -= 14
    c.drawString(0.75 * inch, y, f"Roth Deferral Rate: {r['roth_deferral_rate']}%")

    # Holding
    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(vg_red)
    c.drawString(0.75 * inch, y, "Investment Summary")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=vg_red, width=1)
    y -= 14

    h = r["holdings"][0]
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"{h['name']} ({h['ticker']})")
    c.drawRightString(5 * inch, y, _fmt(h["balance"]))
    c.drawRightString(6 * inch, y, f"{h['allocation']}%")
    c.drawRightString(7.5 * inch, y, _fmt(h["gain_loss"]))

    _draw_footer(c, "The Vanguard Group, Inc.")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_401k_troweprice():
    """T. Rowe Price quarterly statement with multiple funds."""
    r = RETIREMENT_TROWEPRICE
    filepath = OUTPUT_DIR / "retirement" / "401k-troweprice-quarterly.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)
    trp_blue = HexColor("#003B71")

    _draw_branded_header(c, "T. Rowe Price", "Quarterly Account Statement", color=trp_blue)

    y = 10.0 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Plan: {r['plan_name']}")
    c.drawString(0.75 * inch, y - 12, f"Participant: {PERSONA['name']}")
    c.drawString(4.5 * inch, y, f"Statement Period: {r['statement_start']} to {r['statement_end']}")
    y -= 35

    y = _generate_401k_common(c, r, y, trp_blue)

    # Balance by type
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(trp_blue)
    c.drawString(0.75 * inch, y, "Activity by Contribution Type")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=trp_blue, width=1)
    y -= 14

    for label, val in [
        ("Roth Balance", _fmt(r["roth_balance"])),
        ("Salary Deferral (Pretax) Balance", _fmt(r["pretax_balance"])),
        ("Employer Match Balance", _fmt(r["employer_match_balance"])),
        ("Total Employee Contributions (since provider)", _fmt(r["employee_contributions_lifetime"])),
        ("Total Employer Contributions (since provider)", _fmt(r["employer_contributions_lifetime"])),
        ("Roth Basis", _fmt(r["roth_basis"])),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.drawRightString(5.5 * inch, y, val)
        y -= 14

    # Rates
    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(trp_blue)
    c.drawString(0.75 * inch, y, "Deferral Rates per Pay Period")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=trp_blue, width=1)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"Salary Deferral: {r['pretax_deferral_rate']}%")
    y -= 14
    c.drawString(0.75 * inch, y, f"Roth Contributions: {r['roth_deferral_rate']}%")

    # Holdings
    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(trp_blue)
    c.drawString(0.75 * inch, y, "Investment Activity")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=trp_blue, width=1)
    y -= 14

    c.setFont("Helvetica-Bold", 7)
    c.drawString(0.75 * inch, y, "Fund Name")
    c.drawString(3.5 * inch, y, "Ticker")
    c.drawRightString(5 * inch, y, "Balance")
    c.drawRightString(6 * inch, y, "Alloc %")
    c.drawRightString(7.5 * inch, y, "Gain/Loss")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    c.setFont("Helvetica", 8)
    for h in r["holdings"]:
        c.drawString(0.75 * inch, y, h["name"])
        c.drawString(3.5 * inch, y, h["ticker"])
        c.drawRightString(5 * inch, y, _fmt(h["balance"]))
        c.drawRightString(6 * inch, y, f"{h['allocation']}%")
        c.drawRightString(7.5 * inch, y, _fmt(h["gain_loss"]))
        y -= 14

    _draw_footer(c, "T. Rowe Price Investment Services, Inc.")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_401k_minimal():
    """Minimal 401k statement -- just total balance and one fund."""
    r = RETIREMENT_MINIMAL
    filepath = OUTPUT_DIR / "retirement" / "401k-minimal.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(0.75 * inch, 10.3 * inch, "401(k) Plan Statement")
    c.setFont("Helvetica", 10)
    c.drawString(0.75 * inch, 10.1 * inch, f"{r['provider']}")

    y = 9.8 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Plan: {r['plan_name']}")
    c.drawString(0.75 * inch, y - 14, f"Participant: {PERSONA['name']}")
    y -= 40

    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Account Balance")
    y -= 5
    _draw_hr(c, y)
    y -= 18

    c.setFont("Helvetica-Bold", 14)
    c.drawString(0.75 * inch, y, f"Total Balance: {_fmt(r['total_balance'])}")
    y -= 30

    c.setFont("Helvetica", 8)
    h = r["holdings"][0]
    c.drawString(0.75 * inch, y, f"Investment: {h['name']} ({h['ticker']})")
    c.drawString(0.75 * inch, y - 14, f"Allocation: {h['allocation']}%")

    _draw_footer(c, r["provider"])
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_401k_with_loan():
    """401k statement showing an outstanding loan."""
    r = RETIREMENT_WITH_LOAN
    filepath = OUTPUT_DIR / "retirement" / "401k-with-loan.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)
    fid_green = HexColor("#4A8C2A")

    _draw_branded_header(c, "Fidelity Investments", "Quarterly Statement", color=fid_green)

    y = 10.0 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Plan: {r['plan_name']}")
    c.drawString(0.75 * inch, y - 12, f"Participant: {PERSONA['name']}")
    c.drawString(4.5 * inch, y, f"Statement Period: {r['statement_start']} to {r['statement_end']}")
    y -= 40

    y = _generate_401k_common(c, r, y, fid_green)

    # Balance by type
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Balance by Contribution Type")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14

    for label, val in [
        ("Roth Balance", _fmt(r["roth_balance"])),
        ("Salary Deferral (Pre-Tax)", _fmt(r["pretax_balance"])),
        ("Employer Match", _fmt(r["employer_match_balance"])),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.drawRightString(5.5 * inch, y, val)
        y -= 14

    # Loan section
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(HexColor("#CC0000"))
    c.drawString(0.75 * inch, y, "Outstanding Loan")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#CC0000"), width=1)
    y -= 14

    loan_fields = [
        ("Loan Balance", _fmt(r["loan_balance"])),
        ("Loan Interest Rate", _pct(r["loan_rate"])),
        ("Monthly Payment", _fmt(r["loan_payment"])),
        ("Loan Origination Date", r["loan_origination"]),
    ]
    for label, val in loan_fields:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.drawRightString(5.5 * inch, y, val)
        y -= 14

    # Contribution rates
    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Current Contribution Rates")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"Pre-Tax Deferral: {r['pretax_deferral_rate']}%")
    y -= 14
    c.drawString(0.75 * inch, y, f"Roth Deferral: {r['roth_deferral_rate']}%")

    # Holdings
    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(fid_green)
    c.drawString(0.75 * inch, y, "Investment Holdings")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=fid_green, width=1)
    y -= 14

    h = r["holdings"][0]
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, f"{h['name']} ({h['ticker']})")
    c.drawRightString(5 * inch, y, _fmt(h["balance"]))
    c.drawRightString(6 * inch, y, f"{h['allocation']}%")
    c.drawRightString(7.5 * inch, y, _fmt(h["gain_loss"]))

    _draw_footer(c, "Fidelity Investments")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


# =============================================================================
# PROPERTY / MORTGAGE GENERATORS
# =============================================================================

def generate_mortgage_chase():
    """Chase mortgage statement."""
    m = MORTGAGE
    filepath = OUTPUT_DIR / "property" / "mortgage-chase.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    _draw_branded_header(c, "Chase", "Mortgage Statement", color=HexColor("#117ACA"))

    y = 10.0 * inch

    # Account info
    _draw_box(c, 0.75 * inch, y - 0.6 * inch, 6.75 * inch, 0.6 * inch, fill=HexColor("#F0F6FC"), stroke=HexColor("#117ACA"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.9 * inch, y - 0.15 * inch, f"Account Number: {m['account_number']}")
    c.drawString(4 * inch, y - 0.15 * inch, f"Statement Date: 03/01/2026")
    c.setFont("Helvetica", 8)
    c.drawString(0.9 * inch, y - 0.3 * inch, f"Borrower: {PERSONA['name']}")
    c.drawString(4 * inch, y - 0.3 * inch, f"Payment Due Date: 04/01/2026")
    c.drawString(0.9 * inch, y - 0.45 * inch, f"Property: {m['property_address']}")
    y -= 0.8 * inch

    # Mortgage info
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(HexColor("#117ACA"))
    c.drawString(0.75 * inch, y, "Loan Information")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#117ACA"), width=1)
    y -= 14

    info_fields = [
        ("Original Principal Balance", _fmt(m["original_amount"])),
        ("Unpaid Principal Balance", _fmt(m["current_balance"])),
        ("Interest Rate", f"{m['rate']}%"),
        ("Loan Origination Date", m["origination_date"]),
        ("Maturity Date", m["maturity_date"]),
        ("Escrow Balance", _fmt(m["escrow_balance"])),
    ]

    for label, val in info_fields:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    # Payment breakdown
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(HexColor("#117ACA"))
    c.drawString(0.75 * inch, y, "Explanation of Amount Due")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#117ACA"), width=1)
    y -= 14

    payment_fields = [
        ("Principal", _fmt(m["principal"])),
        ("Interest", _fmt(m["interest"])),
        ("Escrow", _fmt(m["escrow"])),
    ]
    for label, val in payment_fields:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    _draw_hr(c, y + 5, color=black)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.75 * inch, y - 5, "Total Payment Due")
    c.setFont("Courier-Bold", 10)
    c.drawRightString(5.5 * inch, y - 5, _fmt(m["monthly_payment"]))

    # Transaction activity (visual noise)
    y -= 35
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(HexColor("#117ACA"))
    c.drawString(0.75 * inch, y, "Recent Activity")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=HexColor("#117ACA"), width=1)
    y -= 14
    c.setFont("Helvetica", 7)
    activity = [
        ("03/01/2026", "Payment Received - Thank You", "", _fmt(m["monthly_payment"])),
        ("02/01/2026", "Payment Received - Thank You", "", _fmt(m["monthly_payment"])),
        ("01/01/2026", "Payment Received - Thank You", "", _fmt(m["monthly_payment"])),
    ]
    for dt, desc, debit, credit in activity:
        c.drawString(0.75 * inch, y, dt)
        c.drawString(2 * inch, y, desc)
        c.drawRightString(7.5 * inch, y, credit)
        y -= 12

    # Contact info (noise)
    y -= 20
    _draw_box(c, 0.75 * inch, y - 0.5 * inch, 6.75 * inch, 0.5 * inch, fill=HexColor("#F0F6FC"))
    c.setFont("Helvetica", 7)
    c.drawString(0.9 * inch, y - 0.15 * inch, "Questions about your mortgage? Visit chase.com/mortgage or call 1-800-848-9136.")
    c.drawString(0.9 * inch, y - 0.3 * inch, "For payment options, visit chase.com/mortgagepayments.")

    _draw_footer(c, "JPMorgan Chase Bank, N.A.")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_mortgage_wells_fargo():
    """Wells Fargo mortgage statement with escrow analysis."""
    m = MORTGAGE
    filepath = OUTPUT_DIR / "property" / "mortgage-wells-fargo.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)
    wf_red = HexColor("#CC0000")

    _draw_branded_header(c, "Wells Fargo Home Mortgage", "Monthly Mortgage Statement", color=wf_red)

    y = 10.0 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Loan Number: {m['account_number']}")
    c.drawString(0.75 * inch, y - 14, f"Borrower: {PERSONA['name']}")
    c.drawString(0.75 * inch, y - 28, f"Property Address: {m['property_address']}")
    c.drawString(4.5 * inch, y, f"Statement Date: March 1, 2026")
    c.drawString(4.5 * inch, y - 14, f"Amount Due: {_fmt(m['monthly_payment'])}")
    c.drawString(4.5 * inch, y - 28, f"Due Date: April 1, 2026")
    y -= 50

    # Loan details
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(wf_red)
    c.drawString(0.75 * inch, y, "Loan Details")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=wf_red, width=1)
    y -= 14

    for label, val in [
        ("Original Loan Amount", _fmt(m["original_amount"])),
        ("Current Principal Balance", _fmt(m["current_balance"])),
        ("Interest Rate", f"{m['rate']}%"),
        ("Loan Type", "30-Year Fixed"),
        ("Origination Date", m["origination_date"]),
        ("Maturity Date", m["maturity_date"]),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    # Payment breakdown
    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(wf_red)
    c.drawString(0.75 * inch, y, "Payment Breakdown")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=wf_red, width=1)
    y -= 14

    for label, val in [
        ("Principal", _fmt(m["principal"])),
        ("Interest", _fmt(m["interest"])),
        ("Escrow (Tax & Insurance)", _fmt(m["escrow"])),
        ("Total Monthly Payment", _fmt(m["monthly_payment"])),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        bold = label.startswith("Total")
        c.setFont("Courier-Bold" if bold else "Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    # Escrow analysis (visual noise / extra data)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(wf_red)
    c.drawString(0.75 * inch, y, "Escrow Account Analysis")
    c.setFillColor(black)
    y -= 5
    _draw_hr(c, y, color=wf_red, width=1)
    y -= 14

    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, "Current Escrow Balance")
    c.drawRightString(5.5 * inch, y, _fmt(m["escrow_balance"]))
    y -= 14
    c.drawString(0.75 * inch, y, "Annual Property Tax (est.)")
    c.drawRightString(5.5 * inch, y, "$5,850.00")
    y -= 14
    c.drawString(0.75 * inch, y, "Annual Homeowners Insurance (est.)")
    c.drawRightString(5.5 * inch, y, "$1,920.00")
    y -= 14
    c.drawString(0.75 * inch, y, "Monthly Escrow Payment")
    c.drawRightString(5.5 * inch, y, _fmt(m["escrow"]))

    # Legal notice
    y -= 30
    c.setFont("Helvetica", 6)
    c.setFillColor(grey)
    c.drawString(0.75 * inch, y, "IMPORTANT NOTICES: If your escrow account has a shortage or surplus, your monthly payment may change.")
    c.drawString(0.75 * inch, y - 8, "Wells Fargo Home Mortgage is a division of Wells Fargo Bank, N.A. NMLSR ID 399801.")
    c.setFillColor(black)

    _draw_footer(c, "Wells Fargo Home Mortgage")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_mortgage_refi():
    """Refinance closing statement with new terms."""
    filepath = OUTPUT_DIR / "property" / "mortgage-refi-statement.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(4.25 * inch, 10.3 * inch, "Closing Disclosure")
    c.setFont("Helvetica", 9)
    c.drawCentredString(4.25 * inch, 10.1 * inch, "This form is a statement of final loan terms and closing costs.")
    _draw_hr(c, 10.0 * inch, color=black, width=1.5)

    y = 9.7 * inch
    c.setFont("Helvetica", 8)

    # Transaction info
    c.drawString(0.75 * inch, y, f"Date Issued: 01/15/2026")
    c.drawString(4.5 * inch, y, f"Closing Date: 01/30/2026")
    y -= 14
    c.drawString(0.75 * inch, y, f"Borrower: {PERSONA['name']}")
    c.drawString(4.5 * inch, y, f"Property: {MORTGAGE['property_address']}")
    y -= 14
    c.drawString(0.75 * inch, y, f"Lender: First National Refinance Corp")
    c.drawString(4.5 * inch, y, f"Loan Purpose: Refinance")
    y -= 30

    # Loan Terms
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "LOAN TERMS")
    y -= 5
    _draw_hr(c, y, color=black, width=1)
    y -= 18

    terms = [
        ("Loan Amount", "$355,000.00"),
        ("Interest Rate", "3.625%"),
        ("Monthly Principal & Interest", "$1,618.72"),
        ("Estimated Total Monthly Payment", "$1,985.00"),
        ("Estimated Escrow", "$366.28"),
        ("Loan Term", "30 years"),
        ("Loan Product", "Fixed Rate"),
    ]
    for label, val in terms:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        c.setFont("Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    # Closing costs summary (visual noise)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "CLOSING COST DETAILS")
    y -= 5
    _draw_hr(c, y, color=black, width=1)
    y -= 18

    costs = [
        ("Origination Fee", "$1,775.00"),
        ("Appraisal Fee", "$450.00"),
        ("Credit Report", "$30.00"),
        ("Title Insurance", "$1,250.00"),
        ("Recording Fees", "$125.00"),
        ("Government Recording", "$85.00"),
        ("Total Closing Costs", "$3,715.00"),
    ]
    for label, val in costs:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        bold = label.startswith("Total")
        c.setFont("Courier-Bold" if bold else "Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    c.setFont("Helvetica", 7)
    c.setFillColor(grey)
    c.drawString(0.75 * inch, y - 15, "This Closing Disclosure is provided in accordance with the Real Estate Settlement Procedures Act (RESPA).")
    c.setFillColor(black)

    _draw_footer(c, "First National Refinance Corp")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_property_tax_bill():
    """County property tax bill -- tests non-mortgage property doc handling."""
    filepath = OUTPUT_DIR / "property" / "property-tax-bill.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    _draw_branded_header(c, "Hampden County", "Real Estate Tax Bill - FY2026", color=HexColor("#2C5F2D"))

    y = 10.0 * inch
    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, f"Bill Number: T-2026-045832")
    c.drawString(4.5 * inch, y, f"Due Date: 02/01/2026")
    y -= 14
    c.drawString(0.75 * inch, y, f"Owner: {PERSONA['name']}")
    c.drawString(4.5 * inch, y, f"Map-Lot: 042-0128-0000")
    y -= 14
    c.drawString(0.75 * inch, y, f"Property: {MORTGAGE['property_address']}")
    y -= 30

    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Tax Assessment")
    y -= 5
    _draw_hr(c, y)
    y -= 14

    for label, val in [
        ("Assessed Land Value", "$145,000.00"),
        ("Assessed Building Value", "$375,000.00"),
        ("Total Assessed Value", "$520,000.00"),
        ("Tax Rate (per $1,000)", "$11.25"),
        ("Total Tax", "$5,850.00"),
        ("First Half (due 02/01)", "$2,925.00"),
        ("Second Half (due 05/01)", "$2,925.00"),
    ]:
        c.setFont("Helvetica", 8)
        c.drawString(0.75 * inch, y, label)
        bold = "Total" in label
        c.setFont("Courier-Bold" if bold else "Courier", 9)
        c.drawRightString(5.5 * inch, y, val)
        y -= 15

    # Exemptions (noise)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Exemptions Applied")
    y -= 5
    _draw_hr(c, y)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(0.75 * inch, y, "Residential Exemption: None")
    y -= 14
    c.drawString(0.75 * inch, y, "Veteran's Exemption: None")

    _draw_footer(c, "Springfield City Collector's Office")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


# =============================================================================
# RSU GENERATORS
# =============================================================================

def generate_rsu_etrade():
    """E-Trade RSU vesting schedule Excel file."""
    filepath = OUTPUT_DIR / "rsu" / "rsu-etrade-vesting.xlsx"
    wb = openpyxl.Workbook()

    # Remove default sheet, create Restricted Stock sheet
    wb.remove(wb.active)
    ws = wb.create_sheet("Restricted Stock")

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # E-Trade branding header rows
    ws.merge_cells("A1:H1")
    ws["A1"] = "E*TRADE Securities LLC"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="003366")
    ws.merge_cells("A2:H2")
    ws["A2"] = "Restricted Stock Unit Activity Report"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws.merge_cells("A3:H3")
    ws["A3"] = f"Account: {PERSONA['name']}  |  As of: 03/15/2026"
    ws["A3"].font = Font(name="Calibri", size=10, color="333333")

    row = 5

    for grant in RSU_GRANTS:
        # Grant header row
        grant_headers = ["Row Type", "Grant Number", "Symbol", "Grant Date",
                         "Granted Qty.", "Vested Qty.", "Unvested Qty.",
                         "Sellable Qty.", "Est. Market Value"]
        for col_idx, header in enumerate(grant_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Grant data row
        grant_data = [
            "Grant", grant["grant_number"], grant["symbol"],
            grant["grant_date"], grant["total_shares"],
            grant["vested_shares"], grant["unvested_shares"],
            grant["sellable_shares"], grant["market_value"],
        ]
        for col_idx, val in enumerate(grant_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if col_idx == 9:
                cell.number_format = money_fmt
        row += 1

        # Vest Schedule header
        vest_headers = ["Row Type", "Grant Number", "Vest Period", "Vest Date",
                        "Vested Qty.", "Released Qty.", "Total Taxes Paid",
                        "Shares Traded for taxes", ""]
        for col_idx, header in enumerate(vest_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="003366")
            cell.fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
            cell.border = border
        row += 1

        # Vest schedule data rows
        for ve in grant["vest_events"]:
            vest_data = [
                "Vest Schedule", grant["grant_number"], ve["period"],
                ve["date"], ve["shares"], ve["released"],
                ve["taxes_paid"], ve["shares_traded"], "",
            ]
            for col_idx, val in enumerate(vest_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = border
                if col_idx == 7:
                    cell.number_format = money_fmt
            row += 1

        # Tax Withholding header
        tax_headers = ["Row Type", "Grant Number", "Vest Period",
                       "Tax Description", "Taxable Gain",
                       "Effective Tax Rate", "Withholding Amount", "", ""]
        for col_idx, header in enumerate(tax_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="003366")
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            cell.border = border
        row += 1

        # Tax data rows
        for ve in grant["vest_events"]:
            if ve["taxes_paid"] > 0:
                # Federal
                tax_data_fed = [
                    "Tax Withholding", grant["grant_number"], ve["period"],
                    "Federal Income Tax", ve["taxable_gain"],
                    ve["eff_rate"] / 100 * 0.7, ve["taxes_paid"] * 0.55, "", "",
                ]
                for col_idx, val in enumerate(tax_data_fed, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = border
                    if col_idx in (5, 7):
                        cell.number_format = money_fmt
                    if col_idx == 6:
                        cell.number_format = pct_fmt
                row += 1

                # State
                tax_data_state = [
                    "Tax Withholding", grant["grant_number"], ve["period"],
                    "State Income Tax", ve["taxable_gain"],
                    ve["eff_rate"] / 100 * 0.15, ve["taxes_paid"] * 0.2, "", "",
                ]
                for col_idx, val in enumerate(tax_data_state, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = border
                    if col_idx in (5, 7):
                        cell.number_format = money_fmt
                    if col_idx == 6:
                        cell.number_format = pct_fmt
                row += 1

                # Social Security
                tax_data_ss = [
                    "Tax Withholding", grant["grant_number"], ve["period"],
                    "Social Security Tax", ve["taxable_gain"],
                    ve["eff_rate"] / 100 * 0.1, ve["taxes_paid"] * 0.15, "", "",
                ]
                for col_idx, val in enumerate(tax_data_ss, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = border
                    if col_idx in (5, 7):
                        cell.number_format = money_fmt
                    if col_idx == 6:
                        cell.number_format = pct_fmt
                row += 1

                # Medicare
                tax_data_med = [
                    "Tax Withholding", grant["grant_number"], ve["period"],
                    "Medicare Tax", ve["taxable_gain"],
                    ve["eff_rate"] / 100 * 0.05, ve["taxes_paid"] * 0.1, "", "",
                ]
                for col_idx, val in enumerate(tax_data_med, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = border
                    if col_idx in (5, 7):
                        cell.number_format = money_fmt
                    if col_idx == 6:
                        cell.number_format = pct_fmt
                row += 1

        row += 1  # blank row between grants

    # Set column widths
    widths = [14, 14, 10, 14, 12, 12, 16, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Disclaimer at bottom
    row += 1
    ws.merge_cells(f"A{row}:I{row}")
    ws.cell(row=row, column=1, value="This report is provided for informational purposes only. E*TRADE Securities LLC, Member SIPC.").font = Font(size=8, color="999999")

    wb.save(str(filepath))
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_rsu_schwab():
    """Charles Schwab RSU vesting schedule -- different column layout."""
    filepath = OUTPUT_DIR / "rsu" / "rsu-schwab-vesting.xlsx"
    wb = openpyxl.Workbook()

    wb.remove(wb.active)
    ws = wb.create_sheet("Stock Plan Activity")

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00539B", end_color="00539B", fill_type="solid")
    data_font = Font(name="Arial", size=9)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Branding
    ws.merge_cells("A1:J1")
    ws["A1"] = "Charles Schwab"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="00539B")
    ws.merge_cells("A2:J2")
    ws["A2"] = "Equity Award Center - RSU Vesting Report"
    ws["A2"].font = Font(name="Arial", size=11, color="666666")
    ws.merge_cells("A3:J3")
    ws["A3"] = f"Participant: {PERSONA['name']}  |  Company: Microsoft Corporation  |  Report Date: 03/15/2026"
    ws["A3"].font = Font(name="Arial", size=9, color="333333")

    # Headers -- Schwab uses a flat layout (different from E-Trade)
    row = 5
    headers = [
        "Award ID", "Award Type", "Symbol", "Award Date",
        "Vest Date", "Status", "Shares Awarded", "Shares Vested",
        "FMV at Vest", "Tax Withholding ($)",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Flatten grants into rows (Schwab's format)
    for grant in RSU_GRANTS:
        for ve in grant["vest_events"]:
            is_vested = ve["shares"] > 0 and ve["released"] > 0
            data = [
                grant["grant_number"], "RSU", grant["symbol"],
                grant["grant_date"], ve["date"],
                "Vested" if is_vested else "Pending",
                grant["total_shares"], ve["shares"],
                ve["taxable_gain"] / ve["shares"] if ve["taxable_gain"] > 0 else "",
                ve["taxes_paid"] if ve["taxes_paid"] > 0 else "",
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = border
                if col_idx in (9, 10) and isinstance(val, (int, float)) and val:
                    cell.number_format = '#,##0.00'
            row += 1

    # Column widths
    widths = [12, 10, 8, 12, 12, 10, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Disclaimer
    row += 2
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row=row, column=1, value="Schwab Stock Plan Services provides equity compensation plan services and administration. Charles Schwab & Co., Inc., Member SIPC.").font = Font(size=7, color="999999")

    wb.save(str(filepath))
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


def generate_rsu_grant_letter():
    """PDF grant confirmation letter for RSU."""
    filepath = OUTPUT_DIR / "rsu" / "rsu-simple-grant-letter.pdf"
    c = canvas.Canvas(str(filepath), pagesize=letter)

    # Microsoft header
    _draw_branded_header(c, "Microsoft Corporation", "Stock Award Notification", color=HexColor("#00A4EF"))

    y = 10.0 * inch
    grant = RSU_GRANTS[1]  # Use the 2024 grant

    # Letter content
    c.setFont("Helvetica", 10)
    c.drawString(0.75 * inch, y, f"Date: May 15, 2024")
    y -= 20
    c.drawString(0.75 * inch, y, f"To: {PERSONA['first']} {PERSONA['last']}")
    c.drawString(0.75 * inch, y - 14, f"Employee ID: E-482910")
    c.drawString(0.75 * inch, y - 28, f"Department: Cloud & AI")
    y -= 55

    c.setFont("Helvetica", 9)
    c.drawString(0.75 * inch, y, "Dear Claudius,")
    y -= 20
    c.drawString(0.75 * inch, y, "We are pleased to inform you that you have been granted the following Restricted Stock Unit (RSU) award")
    y -= 12
    c.drawString(0.75 * inch, y, "under the Microsoft Corporation 2017 Stock Plan:")
    y -= 30

    # Grant details box
    _draw_box(c, 0.75 * inch, y - 1.2 * inch, 6.75 * inch, 1.2 * inch, fill=HexColor("#F0F8FF"), stroke=HexColor("#00A4EF"))

    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.9 * inch, y - 0.15 * inch, "RSU Grant Details")
    c.setFont("Helvetica", 9)

    details = [
        ("Grant Number", grant["grant_number"]),
        ("Stock Symbol", grant["symbol"]),
        ("Grant Date", grant["grant_date"]),
        ("Total Shares Granted", str(grant["total_shares"])),
        ("Vesting Schedule", "4 equal installments over 2 years (semi-annual)"),
    ]
    detail_y = y - 0.35 * inch
    for label, val in details:
        c.setFont("Helvetica-Bold", 8)
        c.drawString(0.9 * inch, detail_y, f"{label}:")
        c.setFont("Helvetica", 8)
        c.drawString(2.8 * inch, detail_y, val)
        detail_y -= 14

    y -= 1.5 * inch

    # Vesting schedule
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.75 * inch, y, "Vesting Schedule")
    y -= 5
    _draw_hr(c, y)
    y -= 14

    c.setFont("Helvetica-Bold", 8)
    c.drawString(0.75 * inch, y, "Period")
    c.drawString(2 * inch, y, "Vest Date")
    c.drawString(3.5 * inch, y, "Shares")
    c.drawString(5 * inch, y, "Status")
    y -= 4
    _draw_hr(c, y)
    y -= 12

    c.setFont("Helvetica", 8)
    for ve in grant["vest_events"]:
        status = "Vested" if ve["released"] > 0 else "Pending"
        c.drawString(0.75 * inch, y, str(ve["period"]))
        c.drawString(2 * inch, y, ve["date"])
        c.drawString(3.5 * inch, y, str(ve["shares"]))
        c.drawString(5 * inch, y, status)
        y -= 14

    # Legal text
    y -= 20
    c.setFont("Helvetica", 7)
    c.setFillColor(grey)
    lines = [
        "This award is subject to the terms and conditions of the Microsoft Corporation 2017 Stock Plan and the",
        "Restricted Stock Unit Award Agreement. Shares will be delivered to your E*Trade account within 5 business",
        "days of each vesting date. Tax withholding will be applied as required by applicable law.",
        "",
        "If you have questions about this award, please contact the Microsoft Stock Administration team at",
        "stockadmin@microsoft.com or visit https://stockplan.microsoft.com.",
        "",
        "This notification does not constitute a contract of employment. Your employment remains at-will.",
    ]
    for line in lines:
        c.drawString(0.75 * inch, y, line)
        y -= 10
    c.setFillColor(black)

    _draw_footer(c, "Microsoft Corporation Stock Administration")
    c.save()
    print(f"  Created: {filepath.relative_to(OUTPUT_DIR)}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    # Create output directories
    for subdir in ["income", "retirement", "property", "rsu"]:
        (OUTPUT_DIR / subdir).mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("Generating Claudius Banks comprehensive test documents...")
    print("=" * 70)
    print()

    print("Income Documents (6 PDFs):")
    generate_w2_standard()
    generate_w2_minimal()
    generate_w2_multiple_states()
    generate_w2_high_earner()
    generate_paystub_adp()
    generate_paystub_workday()
    print()

    print("Retirement Documents (5 PDFs):")
    generate_401k_fidelity()
    generate_401k_vanguard()
    generate_401k_troweprice()
    generate_401k_minimal()
    generate_401k_with_loan()
    print()

    print("Property Documents (4 PDFs):")
    generate_mortgage_chase()
    generate_mortgage_wells_fargo()
    generate_mortgage_refi()
    generate_property_tax_bill()
    print()

    print("RSU Documents (3 files):")
    generate_rsu_etrade()
    generate_rsu_schwab()
    generate_rsu_grant_letter()
    print()

    print("=" * 70)
    total = sum(1 for _ in OUTPUT_DIR.rglob("*") if _.is_file() and _.suffix in (".pdf", ".xlsx"))
    print(f"Generated {total} test documents in: {OUTPUT_DIR}")
    print("=" * 70)
